from datetime import date, datetime, time
import sys
from types import SimpleNamespace

from django.contrib.auth.models import User
from django.db.models import Max
from django.test import RequestFactory, TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
import numpy as np
import pandas as pd
import tempfile
from pathlib import Path
from unittest.mock import patch

from dashboard.models import ForecastSchedule
from dashboard.services.forecast_engine import (
    _station_data_shift_hours as forecast_station_shift_hours,
    _station_capacity_mw as forecast_station_capacity_mw,
    _station_model_dir as forecast_station_model_dir,
    _target_offsets_for_weekday_calendar,
    _postprocess_xgb_prediction,
    _xgb_is_systematically_low,
    _compute_features,
    _apply_early_morning_history_cap,
    _apply_single_axis_tracker_postprocessing,
    _apply_tracker_midday_expected_floor,
    _is_single_axis_tracker,
    WeatherFetchResult,
    run_forecast_for_station,
)
from dashboard.services.forecast_diagnostics import compare_forecast_modes
from dashboard.services.forecast_guardrails import (
    apply_visual_crossing_fallback,
    fallback_heuristic_1_2mw,
    is_bad_irradiation,
)
from dashboard.services.forecast_scheduler import _normalize_schedule_providers, run_scheduled_forecasts
from dashboard.services.model_storage import (
    canonical_station_model_dir,
    cleanup_orphan_model_artifacts,
    describe_station_model_dir,
    find_any_legacy_station_model_dir,
    legacy_root_model_paths,
    legacy_station_model_dir,
    normalize_model_cache,
    resolve_station_model_dir,
)
from dashboard.services.train_models import _prepare_xgb_training_frame, _station_model_dir as train_station_model_dir
from dashboard.services.train_models import add_tracker_training_features
from dashboard.services.train_models import _capacity_mw_from_fields

from dashboard.services.tracker_pvlib_predict import add_features as add_tracker_predict_features
from dashboard.services.tracker_pvlib_predict import _apply_morning_shape_correction
from dashboard.services.tracker_pvlib_training import tracker_config_from_station
from dashboard.views import (
    _build_forecast_plan_map,
    _forecast_export_filters,
    _forecast_value_to_kw,
    _parse_history_datetime,
    _plan_value_with_heuristic_fallback,
    station_forecast_scheduler_tick,
)
from dashboard.services.history_autofill import (
    _station_data_shift_hours as auto_history_station_shift_hours,
    _normalize_auto_history_script,
    _resolve_station_share_folder,
    collect_share_history_dataframe,
    run_auto_history_updates,
    upsert_station_history_from_share,
)

from dashboard.forms import ForecastScheduleForm, StationForm
from dashboard.management.commands.run_scheduled_forecasts import _run_auto_history_updates_safe
from solar.models import SolarForecast, SolarRecord
from stations.models import Organization, OrganizationMember, Station


class ForecastGuardrailTests(TestCase):
    def test_bad_irradiation_detection_respects_daytime_only(self):
        self.assertFalse(is_bad_irradiation(0, 5))
        self.assertFalse(is_bad_irradiation(0, 21))
        self.assertTrue(is_bad_irradiation(None, 12))
        self.assertTrue(is_bad_irradiation("", 12))
        self.assertTrue(is_bad_irradiation(0, 12))
        self.assertTrue(is_bad_irradiation(1301, 12))
        self.assertFalse(is_bad_irradiation(700, 12))

    def test_fallback_heuristic_for_summer_noon_uses_1_2mw_profile(self):
        self.assertEqual(fallback_heuristic_1_2mw("2026-05-07 12:00:00"), 1.14)

    def test_apply_visual_crossing_fallback_replaces_only_bad_daytime_irradiation(self):
        df = pd.DataFrame(
            [
                {"timestamp": "2026-05-07 12:00:00", "irradiation": 0, "pred_final_mw": 0.03},
                {"timestamp": "2026-05-07 22:00:00", "irradiation": 0, "pred_final_mw": 0.0},
                {"timestamp": "2026-05-07 13:00:00", "irradiation": 800, "pred_final_mw": 0.5},
            ]
        )
        guarded = apply_visual_crossing_fallback(df)
        self.assertEqual(guarded.loc[0, "guardrail_reason"], "FALLBACK_BAD_IRRADIATION")
        self.assertEqual(guarded.loc[0, "pred_final_raw_mw"], 0.03)
        self.assertEqual(guarded.loc[0, "pred_final_mw"], 1.14)
        self.assertEqual(guarded.loc[1, "guardrail_reason"], "OK")
        self.assertEqual(guarded.loc[1, "pred_final_mw"], 0.0)
        self.assertEqual(guarded.loc[2, "guardrail_reason"], "OK")
        self.assertEqual(guarded.loc[2, "pred_final_mw"], 0.5)


class TrainModelsXgbFramePreparationTests(TestCase):
    def test_preparation_filters_night_zeros_and_keeps_active_rows(self):
        df = pd.DataFrame(
            [
                {"y": 0.0, "Irradiation": 0.0},
                {"y": 0.0, "Irradiation": 10.0},
                {"y": 1000.0, "Irradiation": 120.0},
                {"y": 500.0, "Irradiation": 5.0},
            ]
        )

        out = _prepare_xgb_training_frame(df, cap_mw=50.0)

        self.assertEqual(len(out), 2)
        self.assertIn("y_over_expected", out.columns)
        self.assertTrue((out["y_permw"] > 0).all())


class StationDataShiftHoursTests(TestCase):
    def test_forecast_shift_helper_reads_station_value(self):
        station = SimpleNamespace(data_shift_hours=-1)
        self.assertEqual(forecast_station_shift_hours(station), -1)

    def test_auto_history_shift_helper_fallbacks_to_zero(self):
        station = SimpleNamespace(data_shift_hours="bad")
        self.assertEqual(auto_history_station_shift_hours(station), 0)


class ForecastValueNormalizationTests(TestCase):
    def test_converts_legacy_mw_values_to_kw(self):
        self.assertEqual(_forecast_value_to_kw(8.8, 8.8), 8800.0)
        self.assertEqual(_forecast_value_to_kw(1.2, 1.2), 1200.0)

    def test_keeps_kw_values_as_is(self):
        self.assertEqual(_forecast_value_to_kw(5400.0, 8.8), 5400.0)
        self.assertEqual(_forecast_value_to_kw(350.0, 1.2), 350.0)


class PlanValueFallbackTests(TestCase):
    def test_uses_pred_final_when_it_is_positive(self):
        self.assertEqual(_plan_value_with_heuristic_fallback(123.0, 456.0), 123.0)

    def test_falls_back_to_heuristic_when_final_is_missing_or_zero(self):
        self.assertEqual(_plan_value_with_heuristic_fallback(None, 456.0), 456.0)
        self.assertEqual(_plan_value_with_heuristic_fallback(0.0, 456.0), 456.0)

    def test_build_map_keeps_only_rows_with_plan_value(self):
        rows = [
            {"timestamp": "t1", "pred_final": 100.0, "pred_heur": 90.0},
            {"timestamp": "t2", "pred_final": None, "pred_heur": 80.0},
            {"timestamp": "t3", "pred_final": None, "pred_heur": None},
        ]
        self.assertEqual(
            _build_forecast_plan_map(rows, "timestamp"),
            {"t1": 100.0, "t2": 80.0},
        )


class ForecastExportDateFilterTests(TestCase):
    def test_same_day_range_is_inclusive_for_full_day(self):
        dt = _parse_history_datetime("2026-03-31 00:00")
        filters = _forecast_export_filters(dt, dt, None)
        self.assertEqual(
            filters,
            {
                "timestamp__date__gte": date(2026, 3, 31),
                "timestamp__date__lte": date(2026, 3, 31),
            },
        )

    def test_exact_date_has_priority_over_range(self):
        dt_from = _parse_history_datetime("2026-03-31 00:00")
        dt_to = _parse_history_datetime("2026-04-01 00:00")
        dt_date = _parse_history_datetime("2026-03-31 00:00")
        self.assertEqual(_forecast_export_filters(dt_from, dt_to, dt_date), {"timestamp__date": date(2026, 3, 31)})


class CapacityFieldsNormalizationTests(TestCase):
    def test_train_capacity_helper_converts_kw_in_capacity_mw_field(self):
        station = SimpleNamespace(capacity_mw=8800, capacity_ac_kw=None, capacity_kw=None, capacity_dc_kw=None)
        self.assertEqual(_capacity_mw_from_fields(station), 8.8)

    @patch("dashboard.services.forecast_engine.SolarRecord")
    def test_forecast_capacity_helper_converts_kw_in_capacity_mw_field(self, solar_record_mock):
        qs = solar_record_mock.objects.filter.return_value
        qs.exclude.return_value = qs
        qs.order_by.return_value = qs
        qs.values_list.return_value = []

        station = SimpleNamespace(
            pk=1,
            capacity_mw=8800,
            capacity_ac_kw=None,
            capacity_kw=None,
            capacity_dc_kw=None,
            history_source_id=None,
        )
        self.assertEqual(forecast_station_capacity_mw(station), 8.8)


class StationModelDirResolutionTests(TestCase):
    def test_canonical_model_dir_is_stable_even_if_name_changes(self):
        station = SimpleNamespace(pk=50, name="SES Balkhash")

        first = canonical_station_model_dir(Path("/tmp/models_cache"), station)
        station.name = "SES Balkhash updated"
        second = canonical_station_model_dir(Path("/tmp/models_cache"), station)

        self.assertEqual(first, Path("/tmp/models_cache/50"))
        self.assertEqual(second, first)

    def test_resolve_uses_legacy_slug_dir_when_it_already_exists(self):
        station = SimpleNamespace(pk=50, name="SES Balkhash")

        with tempfile.TemporaryDirectory() as tmpdir:
            model_root = Path(tmpdir)
            legacy_dir = legacy_station_model_dir(model_root, station)
            legacy_dir.mkdir(parents=True, exist_ok=True)

            resolved = resolve_station_model_dir(model_root, station)

        self.assertEqual(resolved, legacy_dir)

    def test_resolve_finds_legacy_dir_with_old_slug_after_station_rename(self):
        station = SimpleNamespace(pk=50, name="SES Balkhash renamed")

        with tempfile.TemporaryDirectory() as tmpdir:
            model_root = Path(tmpdir)
            old_legacy_dir = model_root / "50_ses-balkhash"
            old_legacy_dir.mkdir(parents=True, exist_ok=True)
            (old_legacy_dir / "xgb_model.json").write_text("{}", encoding="utf-8")

            found_legacy = find_any_legacy_station_model_dir(model_root, station)
            resolved = resolve_station_model_dir(model_root, station)

        self.assertEqual(found_legacy, old_legacy_dir)
        self.assertEqual(resolved, old_legacy_dir)

    def test_describe_marks_previous_slug_legacy_source(self):
        station = SimpleNamespace(pk=50, name="SES Balkhash renamed")

        with tempfile.TemporaryDirectory() as tmpdir:
            model_root = Path(tmpdir)
            old_legacy_dir = model_root / "50_ses-balkhash"
            old_legacy_dir.mkdir(parents=True, exist_ok=True)

            resolved, source = describe_station_model_dir(model_root, station)

        self.assertEqual(resolved, old_legacy_dir)
        self.assertEqual(source, "legacy_previous_slug")

    def test_train_and_forecast_share_same_stable_dir(self):
        station = SimpleNamespace(pk=50, name="SES Balkhash")

        with tempfile.TemporaryDirectory() as tmpdir:
            with override_settings(MODEL_DIR=Path(tmpdir)):
                with patch("dashboard.services.train_models.MODEL_DIR", Path(tmpdir)):
                    with patch("dashboard.services.forecast_engine.MODEL_DIR", Path(tmpdir)):
                        train_dir = train_station_model_dir(station)
                        forecast_dir = forecast_station_model_dir(station)

        self.assertEqual(train_dir, Path(tmpdir) / "50")
        self.assertEqual(forecast_dir, train_dir)


    def test_normalize_moves_legacy_dirs_and_root_files_into_canonical_dir(self):
        station = SimpleNamespace(pk=50, name="SES Balkhash")

        with tempfile.TemporaryDirectory() as tmpdir:
            model_root = Path(tmpdir)
            legacy_dir = model_root / "50_old-name"
            legacy_dir.mkdir(parents=True, exist_ok=True)
            (legacy_dir / "xgb_model.json").write_text("{}", encoding="utf-8")
            root_legacy = legacy_root_model_paths(model_root, station)
            root_legacy["legacy_np"].write_text("np", encoding="utf-8")

            result = normalize_model_cache(model_root, [station])

            canonical_dir = model_root / "50"
            self.assertTrue((canonical_dir / "xgb_model.json").exists())
            self.assertTrue((canonical_dir / "np_model.np").exists())
            self.assertFalse(legacy_dir.exists())
            self.assertFalse(root_legacy["legacy_np"].exists())
            self.assertTrue(any("50_old-name/xgb_model.json" in entry for entry in result["moved"]))

    def test_cleanup_removes_orphan_station_dirs_and_legacy_root_files(self):
        active_station = SimpleNamespace(pk=50, name="SES Balkhash")

        with tempfile.TemporaryDirectory() as tmpdir:
            model_root = Path(tmpdir)
            (model_root / "77").mkdir()
            (model_root / "88_old-station").mkdir()
            orphan_root = model_root / "xgb_model_99.json"
            orphan_root.write_text("{}", encoding="utf-8")

            result = cleanup_orphan_model_artifacts(model_root, [active_station])

            self.assertFalse((model_root / "77").exists())
            self.assertFalse((model_root / "88_old-station").exists())
            self.assertFalse(orphan_root.exists())
            self.assertGreaterEqual(len(result["removed"]), 3)


def build_custom_history_dataframe(station):
    now = timezone.now().replace(minute=0, second=0, microsecond=0)
    return pd.DataFrame(
        [
            {
                "ds": pd.Timestamp(now),
                "irradiation": 500.1,
                "air_temp": 20.2,
                "pv_temp": 24.3,
                "power_kw": 700.4,
            }
        ]
    )



class ForecastEngineMorningFeatureTests(TestCase):
    def test_compute_features_adds_soft_morning_ramp_and_does_not_boost_0700(self):
        df = pd.DataFrame(
            [
                {"ds": pd.Timestamp("2026-06-01 07:00:00"), "irradiation": 100.0, "air_temp": 15.0, "wind_speed": 1.0},
                {"ds": pd.Timestamp("2026-06-01 10:00:00"), "irradiation": 500.0, "air_temp": 20.0, "wind_speed": 1.0},
            ]
        )

        out = _compute_features(df, capacity_mw=10.0, lat_deg=47.86)

        self.assertEqual(int(out.iloc[0]["sunrise_hour_flag"]), 1)
        self.assertAlmostEqual(float(out.iloc[0]["solar_ramp_factor"]), 0.65)
        self.assertAlmostEqual(float(out.iloc[0]["Irradiation"]), 100.0)
        self.assertAlmostEqual(float(out.iloc[0]["irradiation_x_ramp"]), 65.0)
        self.assertEqual(int(out.iloc[0]["morning_peak_boost"]), 0)
        self.assertEqual(int(out.iloc[1]["sunrise_hour_flag"]), 0)
        self.assertAlmostEqual(float(out.iloc[1]["solar_ramp_factor"]), 1.0)

    def test_early_morning_history_cap_only_limits_0700_and_0800(self):
        user = User.objects.create_user(username="morning-cap", password="pass")
        org = Organization.objects.create(name="Morning Cap Org", owner=user)
        station = Station.objects.create(
            org=org,
            name="Morning Cap Station",
            capacity_mw=10.0,
            capacity_ac_kw=10000,
            capacity_dc_kw=11000,
        )
        for day in range(1, 7):
            SolarRecord.objects.create(
                station=station,
                timestamp=timezone.datetime(2026, 6, day, 7, 0, tzinfo=timezone.get_current_timezone()),
                history_scope=SolarRecord.HISTORY_SCOPE_MAIN,
                power_kw=1000.0,
                irradiation=100.0,
            )
            SolarRecord.objects.create(
                station=station,
                timestamp=timezone.datetime(2026, 6, day, 8, 0, tzinfo=timezone.get_current_timezone()),
                history_scope=SolarRecord.HISTORY_SCOPE_MAIN,
                power_kw=2000.0,
                irradiation=200.0,
            )

        feat = pd.DataFrame(
            {
                "ds": [
                    pd.Timestamp("2026-06-10 07:00:00"),
                    pd.Timestamp("2026-06-10 08:00:00"),
                    pd.Timestamp("2026-06-10 10:00:00"),
                ]
            }
        )

        with override_settings(FORECAST_EARLY_MORNING_CAP_UPLIFT=1.10, FORECAST_EARLY_MORNING_CAP_MIN_SAMPLES=5):
            capped, caps = _apply_early_morning_history_cap(
                np.array([3.0, 3.0, 3.0]), feat, station, capacity_mw=10.0
            )

        self.assertAlmostEqual(capped[0], 1.1, places=6)
        self.assertAlmostEqual(capped[1], 2.2, places=6)
        self.assertAlmostEqual(capped[2], 3.0, places=6)
        self.assertEqual(set(caps.keys()), {7, 8})


class ForecastEngineXgbPostprocessTests(TestCase):
    def test_converts_per_mw_predictions_to_mw(self):
        raw = [0.1, 0.5, 1.0]
        meta = {"target": "y_per_MW = y / cap_mw", "cap_mw_used": 50.0}

        out = _postprocess_xgb_prediction(raw, meta, capacity_mw=49.0)

        self.assertEqual(out.tolist(), [5.0, 25.0, 50.0])


    def test_converts_over_expected_predictions_to_mw(self):
        raw = [1.0, 0.5]
        meta = {
            "target": "y_over_expected = y / max(y_expected, floor)",
            "cap_mw_used": 50.0,
            "y_expected_floor_mw": 2.5,
        }
        feat = pd.DataFrame({"Irradiation": [600.0, 0.0]})

        out = _postprocess_xgb_prediction(raw, meta, capacity_mw=49.0, df_feat=feat)

        self.assertAlmostEqual(float(out[0]), 27.0, places=2)
        self.assertAlmostEqual(float(out[1]), 1.25, places=2)


    def test_applies_xgb_calibration_multiplier(self):
        raw = [1.0]
        meta = {
            "target": "y_per_MW = y / cap_mw",
            "cap_mw_used": 10.0,
            "xgb_calib_mult": 2.0,
        }

        out = _postprocess_xgb_prediction(raw, meta, capacity_mw=10.0)

        self.assertEqual(out.tolist(), [20.0])

    def test_keeps_legacy_mw_predictions_without_scaling(self):
        raw = [5.0, 10.0]
        meta = {"target": "y_mw"}

        out = _postprocess_xgb_prediction(raw, meta, capacity_mw=50.0)

        self.assertEqual(out.tolist(), [5.0, 10.0])




class ForecastScheduleFormManualSnowFactorTests(TestCase):
    def test_accepts_manual_snow_factor_above_one(self):
        form = ForecastScheduleForm(
            data={
                "enabled": "on",
                "start_at": "",
                "run_time": "06:00",
                "days": 1,
                "horizon_mode": "weekday_calendar",
                "providers": ["visual_crossing"],
                "emails": "",
                "manual_snow_enable": "on",
                "manual_snow_factor": "1.15",
                "manual_snow_dates": "",
            }
        )

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data["manual_snow_factor"], 1.15)

    def test_rejects_manual_snow_factor_above_safe_limit(self):
        form = ForecastScheduleForm(
            data={
                "enabled": "on",
                "start_at": "",
                "run_time": "06:00",
                "days": 1,
                "horizon_mode": "weekday_calendar",
                "providers": ["visual_crossing"],
                "emails": "",
                "manual_snow_enable": "on",
                "manual_snow_factor": "1.6",
                "manual_snow_dates": "",
            }
        )

        self.assertFalse(form.is_valid())
        self.assertIn("manual_snow_factor", form.errors)


class StationFormAutoHistoryFolderInitialTests(TestCase):
    def test_edit_form_shows_station_specific_folder_when_model_has_default_share(self):
        user = User.objects.create_user(username="folder-form", password="pass")
        org = Organization.objects.create(name="Folder Form Org", owner=user)
        station = Station.objects.create(
            org=org,
            name="SES 8.8 MW",
            capacity_mw=8.8,
            auto_history_folder="/mnt/share",
        )

        form = StationForm(instance=station, user=user)

        self.assertEqual(form["auto_history_folder"].value(), f"/mnt/share/org_{org.id}/SES_8.8_MW")



class StationEditAutoHistoryFolderNormalizationTests(TestCase):
    def test_edit_get_normalizes_plain_share_folder_for_existing_station(self):
        user = User.objects.create_user(username="folder-edit", password="pass")
        org = Organization.objects.create(name="Folder Edit Org", owner=user)
        OrganizationMember.objects.create(
            organization=org,
            user=user,
            role=OrganizationMember.ROLE_OWNER,
        )

        station = Station.objects.create(
            org=org,
            name="SES 8.8 MW",
            capacity_mw=8.8,
        )
        Station.objects.filter(pk=station.pk).update(auto_history_folder="/mnt/share")

        self.client.login(username="folder-edit", password="pass")
        response = self.client.get(f"/dashboard/station/{station.pk}/edit/")

        self.assertEqual(response.status_code, 200)
        station.refresh_from_db()
        self.assertEqual(station.auto_history_folder, f"/mnt/share/org_{org.id}/SES_8.8_MW")
    def test_edit_form_prefers_tmp_path_when_org_has_tmp_station(self):
        user = User.objects.create_user(username="folder-form-tmp", password="pass")
        org = Organization.objects.create(name="Folder Form Tmp Org", owner=user)

        Station.objects.create(
            org=org,
            name="SES 1.2 MW",
            capacity_mw=1.2,
            auto_history_folder=f"/tmp/forecast_platform_auto_history/org_{org.id}/SES_1.2_MW",
        )
        station = Station.objects.create(
            org=org,
            name="SES 8.8 MW",
            capacity_mw=8.8,
            auto_history_folder="/mnt/share",
        )

        form = StationForm(instance=station, user=user)

        self.assertEqual(
            form["auto_history_folder"].value(),
            f"/tmp/forecast_platform_auto_history/org_{org.id}/SES_8.8_MW",
        )







class StationFormOrganizationVisibilityTests(TestCase):
    def test_owner_sees_owned_org_in_operator_field(self):
        owner = User.objects.create_user(username="owner-visible", password="pass")
        own_org = Organization.objects.create(name="Owned Org", owner=owner)

        form = StationForm(user=owner)

        org_ids = set(form.fields["org"].queryset.values_list("id", flat=True))
        self.assertIn(own_org.id, org_ids)

    def test_user_does_not_see_other_organizations(self):
        owner = User.objects.create_user(username="owner-only", password="pass")
        other = User.objects.create_user(username="other-only", password="pass")
        own_org = Organization.objects.create(name="Owned Org", owner=owner)
        other_org = Organization.objects.create(name="Hidden Org", owner=other)

        form = StationForm(user=owner)

        org_ids = set(form.fields["org"].queryset.values_list("id", flat=True))
        self.assertIn(own_org.id, org_ids)
        self.assertNotIn(other_org.id, org_ids)

class StationFormAutoHistoryRunTimeParsingTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="time-form", password="pass")
        self.org = Organization.objects.create(name="Time Form Org", owner=self.user)

    def _base_data(self):
        return {
            "name": "SES 1.2 MW",
            "org": self.org.pk,
            "capacity_mw": "1.2",
            "latitude": "47.8",
            "longitude": "67.64",
            "timezone": "Asia/Almaty",
            "capacity_dc_kw": "1274",
            "capacity_ac_kw": "1203",
            "pr_default": "0.88",
            "tilt_deg": "30",
            "azimuth_deg": "180",
            "losses_total_pct": "10",
            "history_source": "",
            "history_scale_by_capacity": "on",
            "auto_history_enabled": "on",
            "auto_history_folder": "/mnt/share/org_1/SES_1.2_MW",
            "auto_history_script": "",
        }

    def test_accepts_ampm_time_from_legacy_browser(self):
        data = self._base_data()
        data["auto_history_run_time"] = "10:55:00 AM"

        form = StationForm(data=data, user=self.user)

        self.assertTrue(form.is_valid(), form.errors.as_json())
        self.assertEqual(form.cleaned_data["auto_history_run_time"], time(10, 55))

    def test_accepts_24h_time(self):
        data = self._base_data()
        data["auto_history_run_time"] = "22:15"

        form = StationForm(data=data, user=self.user)

        self.assertTrue(form.is_valid(), form.errors.as_json())
        self.assertEqual(form.cleaned_data["auto_history_run_time"], time(22, 15))


class StationAutoHistoryFolderFallbackTests(TestCase):
    def test_missing_station_subfolder_uses_alias_folder_with_spaces(self):
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            alias = base / "org_1" / "SES 1.2 MW"
            alias.mkdir(parents=True)
            station = SimpleNamespace(
                auto_history_folder=str(base / "org_1" / "SES_1.2_MW"),
                pk=42,
                org_id=1,
                name="SES 1.2 MW",
            )

            resolved = _resolve_station_share_folder(station, share_root=base)

            self.assertEqual(resolved, alias)

    def test_missing_station_subfolder_falls_back_to_share_root(self):
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            station = SimpleNamespace(auto_history_folder=str(base / "org_1" / "SES_1.2_MW"), pk=42)

            resolved = _resolve_station_share_folder(station, share_root=base)

            self.assertEqual(resolved, base)

    def test_non_share_path_does_not_fallback(self):
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            station = SimpleNamespace(auto_history_folder=str(base / "other" / "path"), pk=43)

            resolved = _resolve_station_share_folder(station, share_root=base / "share")

            self.assertEqual(resolved, Path(station.auto_history_folder))



class StationAutoHistoryCustomScriptTests(TestCase):
    def setUp(self):
        user = User.objects.create_user(username="autohistory-script", password="pass")
        org = Organization.objects.create(name="AutoHistory Script Org", owner=user)
        self.station = Station.objects.create(
            org=org,
            name="Script Station",
            capacity_mw=1.0,
            auto_history_enabled=True,
            auto_history_script="dashboard.tests:build_custom_history_dataframe",
        )

    def test_upsert_uses_station_custom_script(self):
        rows = upsert_station_history_from_share(self.station)

        self.assertEqual(rows, 1)
        rec = SolarRecord.objects.get(station=self.station)
        self.assertAlmostEqual(rec.power_kw, 700.4)

    def test_short_module_name_resolves_from_history_scripts_package(self):
        self.station.auto_history_script = "example_station"

        rows = upsert_station_history_from_share(self.station)

        self.assertEqual(rows, 1)

    def test_file_path_module_can_be_used(self):
        with tempfile.TemporaryDirectory() as tmp:
            fpath = Path(tmp) / "custom_builder.py"
            fpath.write_text(
                "import pandas as pd\n"
                "def build_history_dataframe(station):\n"
                "    return pd.DataFrame([{\n"
                "        'ds': pd.Timestamp('2026-01-01 10:00:00'),\n"
                "        'irradiation': 400.0,\n"
                "        'air_temp': 15.0,\n"
                "        'pv_temp': 20.0,\n"
                "        'power_kw': 500.0,\n"
                "    }])\n"
            )
            self.station.auto_history_script = f"{fpath}:build_history_dataframe"

            rows = upsert_station_history_from_share(self.station)

        self.assertEqual(rows, 1)

    def test_invalid_custom_script_format_falls_back_to_standard_handler(self):
        self.station.auto_history_script = ":build_history_dataframe"

        rows = upsert_station_history_from_share(self.station)

        self.assertEqual(rows, 0)

    def test_script_value_with_help_text_uses_first_token(self):
        self.station.auto_history_script = "example_station или dashboard.services.history_scripts.example_station:build_history_dataframe"

        rows = upsert_station_history_from_share(self.station)

        self.assertEqual(rows, 1)

    def test_script_value_with_spaces_normalizes_to_short_module_name(self):
        self.station.auto_history_script = "example station"

        rows = upsert_station_history_from_share(self.station)

        self.assertEqual(rows, 1)

    def test_script_value_with_dot_in_short_name_normalizes_to_short_module_name(self):
        self.station.auto_history_script = "ses_1.2mw"

        normalized = _normalize_auto_history_script(self.station.auto_history_script)

        self.assertEqual(normalized, "ses_1_2mw")

    def test_script_value_with_dot_in_short_name_runs_history_builder(self):
        self.station.auto_history_script = "ses_1.2mw"

        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            pd.DataFrame(
                [
                    {
                        "ds": "2026-02-26 08:10:00",
                        "Irradiation": 30.7,
                        "Air_Temp": -9.1,
                        "PV_Temp": -9.0,
                        "Power_KW": 220.123,
                    },
                    {
                        "ds": "2026-02-26 08:40:00",
                        "Irradiation": 46.9,
                        "Air_Temp": -9.1,
                        "PV_Temp": -8.9,
                        "Power_KW": 70.222,
                    },
                ]
            ).to_csv(folder / "history_1_2.csv", index=False)
            self.station.auto_history_folder = str(folder)
            self.station.save(update_fields=["auto_history_script", "auto_history_folder"])

            rows = upsert_station_history_from_share(self.station)

        self.assertEqual(rows, 1)
        rec = SolarRecord.objects.get(station=self.station)
        self.assertAlmostEqual(rec.power_kw, 70.22)

    def test_script_value_with_slashes_normalizes_to_short_module_name(self):
        self.station.auto_history_script = "/history_scripts/example_station"

        rows = upsert_station_history_from_share(self.station)

        self.assertEqual(rows, 1)

    def test_script_value_with_dashboard_path_normalizes_to_short_module_name(self):
        self.station.auto_history_script = "dashboard/services/history_scripts/example_station"

        rows = upsert_station_history_from_share(self.station)

        self.assertEqual(rows, 1)

    def test_script_value_with_windows_slashes_normalizes_to_short_module_name(self):
        self.station.auto_history_script = r"history_scripts\example_station"

        rows = upsert_station_history_from_share(self.station)

        self.assertEqual(rows, 1)




class StationAutoHistoryScheduleTests(TestCase):
    def setUp(self):
        user = User.objects.create_user(username="autohistory-time", password="pass")
        org = Organization.objects.create(name="AutoHistory Time Org", owner=user)
        self.station = Station.objects.create(
            org=org,
            name="Timed Script Station",
            capacity_mw=1.0,
            auto_history_enabled=True,
            auto_history_script="dashboard.tests:build_custom_history_dataframe",
            auto_history_run_time=time(6, 0),
        )

    @patch("dashboard.services.history_autofill.timezone.localtime")
    def test_run_auto_history_updates_skips_before_station_time(self, localtime_mock):
        localtime_mock.return_value = timezone.datetime(2026, 2, 20, 5, 30, tzinfo=timezone.get_current_timezone())

        rows = run_auto_history_updates()

        self.assertEqual(rows, 0)

    @patch("dashboard.services.history_autofill.timezone.localtime")
    def test_run_auto_history_updates_runs_once_per_day(self, localtime_mock):
        localtime_mock.return_value = timezone.datetime(2026, 2, 20, 6, 30, tzinfo=timezone.get_current_timezone())

        first_rows = run_auto_history_updates()
        second_rows = run_auto_history_updates()

        self.assertEqual(first_rows, 1)
        self.assertEqual(second_rows, 0)
        self.station.refresh_from_db()
        self.assertEqual(str(self.station.auto_history_last_run_date), "2026-02-20")

    @patch("dashboard.services.history_autofill._safe_upsert_station", return_value=(1, True))
    @patch("dashboard.services.history_autofill.timezone.localtime")
    def test_run_auto_history_updates_allows_near_time_grace_window(self, localtime_mock, upsert_mock):
        self.station.auto_history_last_run_date = None
        self.station.auto_history_run_time = time(9, 20)
        self.station.save(update_fields=["auto_history_last_run_date", "auto_history_run_time"])
        localtime_mock.return_value = timezone.datetime(2026, 2, 24, 9, 19, 31, tzinfo=timezone.get_current_timezone())

        rows = run_auto_history_updates()

        self.assertEqual(rows, 1)
        upsert_mock.assert_called_once()

    @patch("dashboard.services.history_autofill._safe_upsert_station", return_value=(1, True))
    @patch("dashboard.services.history_autofill.timezone.localtime")
    def test_run_auto_history_updates_allows_pre_time_run_when_scheduler_is_sparse(self, localtime_mock, upsert_mock):
        self.station.auto_history_last_run_date = timezone.datetime(2026, 2, 23).date()
        self.station.auto_history_run_time = time(9, 0)
        self.station.save(update_fields=["auto_history_last_run_date", "auto_history_run_time"])
        localtime_mock.return_value = timezone.datetime(2026, 2, 24, 6, 0, tzinfo=timezone.get_current_timezone())

        rows = run_auto_history_updates()

        self.assertEqual(rows, 1)
        upsert_mock.assert_called_once()
        self.station.refresh_from_db()
        self.assertEqual(str(self.station.auto_history_last_run_date), "2026-02-24")


    @patch("dashboard.services.history_autofill._safe_upsert_station", side_effect=[(0, True), (1, True)])
    @patch("dashboard.services.history_autofill.timezone.localtime")
    def test_run_auto_history_updates_retries_same_day_when_no_rows(self, localtime_mock, _safe_upsert_mock):
        localtime_mock.return_value = timezone.datetime(2026, 2, 20, 6, 30, tzinfo=timezone.get_current_timezone())

        first_rows = run_auto_history_updates()
        second_rows = run_auto_history_updates()

        self.assertEqual(first_rows, 0)
        self.assertEqual(second_rows, 1)
        self.station.refresh_from_db()
        self.assertEqual(str(self.station.auto_history_last_run_date), "2026-02-20")

    @patch("dashboard.services.history_autofill._safe_upsert_station", return_value=(1, True))
    @patch("dashboard.services.history_autofill.timezone.localtime")
    def test_run_auto_history_updates_uses_station_timezone(self, localtime_mock, upsert_mock):
        self.station.timezone = "Asia/Almaty"
        self.station.auto_history_last_run_date = None
        self.station.auto_history_run_time = time(8, 0)
        self.station.save(update_fields=["timezone", "auto_history_last_run_date", "auto_history_run_time"])
        localtime_mock.return_value = timezone.datetime(2026, 3, 20, 3, 35, tzinfo=timezone.get_current_timezone())

        rows = run_auto_history_updates()

        self.assertEqual(rows, 1)
        upsert_mock.assert_called_once()
        self.station.refresh_from_db()
        self.assertEqual(str(self.station.auto_history_last_run_date), "2026-03-20")

    @patch("dashboard.services.history_autofill.timezone.localtime")
    def test_run_auto_history_updates_skip_message_uses_station_timezone(self, localtime_mock):
        self.station.timezone = "Asia/Almaty"
        self.station.auto_history_last_run_date = None
        self.station.auto_history_run_time = time(8, 0)
        self.station.save(update_fields=["timezone", "auto_history_last_run_date", "auto_history_run_time"])
        localtime_mock.return_value = timezone.datetime(2026, 3, 20, 2, 59, tzinfo=timezone.get_current_timezone())

        rows = run_auto_history_updates()

        self.assertEqual(rows, 0)
        self.station.refresh_from_db()
        self.assertEqual(self.station.auto_history_last_status, "skipped")
        self.assertIn("station_now=07:59 < run_time=08:00", self.station.auto_history_last_message)


class StationAutoHistoryConfigChangeResetTests(TestCase):
    def test_station_save_resets_last_run_date_when_auto_history_config_changes(self):
        user = User.objects.create_user(username="autohistory-reset", password="pass")
        org = Organization.objects.create(name="AutoHistory Reset Org", owner=user)
        station = Station.objects.create(
            org=org,
            name="Reset Station",
            capacity_mw=1.0,
            auto_history_enabled=True,
            auto_history_run_time=time(6, 0),
            auto_history_last_run_date=timezone.datetime(2026, 2, 20).date(),
        )

        station.auto_history_run_time = time(7, 0)
        station.save()

        station.refresh_from_db()
        self.assertIsNone(station.auto_history_last_run_date)



class StationAutoHistoryMergeSameDateTests(TestCase):
    @patch("dashboard.services.history_autofill.read_meteo_hourly")
    @patch("dashboard.services.history_autofill.read_plant_report_hourly")
    def test_collect_share_history_merges_two_reports_for_same_date(self, plant_mock, meteo_mock):
        meteo_mock.return_value = pd.DataFrame(
            [{"ds": pd.Timestamp("2026-02-17 10:00:00"), "irradiation": 500, "air_temp": 20, "pv_temp": 25}]
        )
        plant_mock.side_effect = [
            pd.DataFrame([{"ds": pd.Timestamp("2026-02-17 10:00:00"), "power_kw": 100.0}]),
            pd.DataFrame([{"ds": pd.Timestamp("2026-02-17 10:00:00"), "power_kw": 30.0}]),
        ]

        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            (folder / "D222152_20260217_0000.csv.gz").write_text("x")
            (folder / "Plant Report_SPP 1.2 MW_17-02-2026_part1.xlsx").write_text("x")
            (folder / "Plant Report_SPP 1.2 MW_17-02-2026_part2.xlsx").write_text("x")

            out = collect_share_history_dataframe(folder)

        self.assertEqual(len(out), 1)
        self.assertAlmostEqual(float(out.iloc[0]["power_kw"]), 130.0)





class StationMountTypeTests(TestCase):
    def test_station_mount_type_defaults_to_fixed_and_form_exposes_label(self):
        user = User.objects.create_user(username="mount-form", password="pass")
        org = Organization.objects.create(name="Mount Org", owner=user)
        station = Station.objects.create(org=org, name="Fixed SES")

        self.assertEqual(station.mount_type, Station.MOUNT_FIXED)

        form = StationForm(instance=station, user=user)
        self.assertIn("mount_type", form.fields)
        self.assertEqual(form.fields["mount_type"].label, "Тип конструкции")
        self.assertEqual(
            list(form.fields["mount_type"].choices),
            [
                (Station.MOUNT_FIXED, "Фиксированный наклон"),
                (Station.MOUNT_SINGLE_AXIS_TRACKER, "Одноосевой трекер"),
                (Station.MOUNT_DUAL_AXIS_TRACKER, "Двухосевой трекер"),
            ],
        )

    def test_single_axis_tracker_detection_is_opt_in_only(self):
        fixed = SimpleNamespace(mount_type=Station.MOUNT_FIXED)
        tracker = SimpleNamespace(mount_type=Station.MOUNT_SINGLE_AXIS_TRACKER)
        tracker_with_dash = SimpleNamespace(mount_type="single-axis-tracker")

        self.assertFalse(_is_single_axis_tracker(fixed))
        self.assertTrue(_is_single_axis_tracker(tracker))
        self.assertTrue(_is_single_axis_tracker(tracker_with_dash))


class SingleAxisTrackerStationConfigTests(TestCase):
    def setUp(self):
        user = User.objects.create_user(username="tracker-station-config", password="pass")
        org = Organization.objects.create(name="Tracker Config Org", owner=user)
        self.station = Station.objects.create(
            org=org,
            name="Tracker 100 MW",
            mount_type=Station.MOUNT_SINGLE_AXIS_TRACKER,
            latitude=43.598,
            longitude=73.761,
            timezone="Asia/Almaty",
            tracker_axis_tilt=0.0,
            tracker_axis_azimuth=0.0,
            tracker_max_angle=60.0,
            tracker_gcr=0.3105,
            tracker_backtrack=True,
            tracker_poa_model="perez",
            tracker_albedo=0.20,
            capacity_mw=84.0,
            capacity_ac_kw=85000,
            capacity_dc_kw=100000,
        )

    def test_tracker_config_uses_station_values_for_real_params(self):
        cfg = tracker_config_from_station(self.station)

        self.assertEqual(cfg.axis_tilt, 0.0)
        self.assertEqual(cfg.axis_azimuth, 0.0)
        self.assertEqual(cfg.max_angle, 60.0)
        self.assertEqual(cfg.gcr, 0.3105)
        self.assertIs(cfg.backtrack, True)
        self.assertEqual(cfg.model, "perez")
        self.assertEqual(cfg.albedo, 0.20)

    def test_predict_pvlib_features_pass_station_tracker_values_to_singleaxis(self):
        captured = {}

        def fake_singleaxis(**kwargs):
            captured.update(kwargs)
            idx = kwargs["apparent_zenith"].index
            return pd.DataFrame(
                {
                    "tracker_theta": np.zeros(len(idx)),
                    "aoi": np.zeros(len(idx)),
                    "surface_tilt": np.full(len(idx), 12.0),
                    "surface_azimuth": np.full(len(idx), 180.0),
                },
                index=idx,
            )

        def fake_get_total_irradiance(**kwargs):
            captured["poa_model"] = kwargs["model"]
            captured["albedo"] = kwargs["albedo"]
            idx = kwargs["ghi"].index
            return pd.DataFrame({"poa_global": np.full(len(idx), 800.0)}, index=idx)

        fake_pvlib = SimpleNamespace(
            solarposition=SimpleNamespace(
                get_solarposition=lambda times, latitude, longitude: pd.DataFrame(
                    {"apparent_zenith": np.full(len(times), 30.0), "azimuth": np.full(len(times), 180.0)},
                    index=times,
                )
            ),
            irradiance=SimpleNamespace(
                erbs=lambda ghi, zenith, times: pd.DataFrame(
                    {"dni": np.full(len(times), 700.0), "dhi": np.full(len(times), 100.0)},
                    index=times,
                ),
                get_extra_radiation=lambda times: pd.Series(np.full(len(times), 1367.0), index=times),
                get_total_irradiance=fake_get_total_irradiance,
            ),
            atmosphere=SimpleNamespace(
                get_relative_airmass=lambda zenith: pd.Series(np.ones(len(zenith)), index=zenith.index)
            ),
            tracking=SimpleNamespace(singleaxis=fake_singleaxis),
        )
        feat = pd.DataFrame({"ds": pd.to_datetime(["2026-06-01 12:00:00"]), "Irradiation": [900.0]})

        with patch.dict(sys.modules, {"pvlib": fake_pvlib}):
            out = add_tracker_predict_features(feat, self.station, capacity_mw=85.0)

        self.assertEqual(captured["axis_tilt"], self.station.tracker_axis_tilt)
        self.assertEqual(captured["axis_azimuth"], self.station.tracker_axis_azimuth)
        self.assertEqual(captured["max_angle"], self.station.tracker_max_angle)
        self.assertEqual(captured["gcr"], self.station.tracker_gcr)
        self.assertEqual(captured["backtrack"], self.station.tracker_backtrack)
        self.assertEqual(captured["poa_model"], self.station.tracker_poa_model)
        self.assertEqual(captured["albedo"], self.station.tracker_albedo)
        self.assertIn("POA_pvlib", out.columns)
        self.assertIn("solar_elevation", out.columns)
        self.assertIn("solar_zenith", out.columns)
        self.assertIn("solar_azimuth", out.columns)
        self.assertIn("aoi", out.columns)
        self.assertIn("is_morning", out.columns)
        self.assertIn("morning_ramp", out.columns)
        self.assertIn("tracker_pvlib_baseline_mw", out.columns)

    def test_morning_shape_correction_lifts_morning_and_preserves_daily_energy(self):
        feat = pd.DataFrame(
            {
                "ds": pd.date_range("2026-06-23 05:00:00", periods=12, freq="h"),
                "POA_pvlib": [80.0, 420.0, 620.0, 700.0, 780.0, 860.0, 900.0, 910.0, 880.0, 760.0, 600.0, 300.0],
                "Irradiation_GHI": [60.0, 310.0, 480.0, 560.0, 650.0, 790.0, 850.0, 870.0, 820.0, 700.0, 520.0, 240.0],
                "solar_elevation": [1.0, 8.0, 16.0, 25.0, 34.0, 45.0, 55.0, 60.0, 56.0, 48.0, 35.0, 18.0],
                "tracker_theta": [55.0, 60.0, 58.0, 48.0, 30.0, 12.0, 0.0, -10.0, -28.0, -45.0, -55.0, -60.0],
            }
        )
        final = np.array([2.0, 17.0, 35.0, 52.0, 70.0, 82.0, 84.0, 84.0, 82.0, 74.0, 60.0, 25.0])
        baseline = np.array([4.0, 32.0, 48.0, 58.0, 68.0, 76.0, 82.0, 83.0, 80.0, 70.0, 55.0, 24.0])
        hist = np.array([8.0, 60.0, 66.0, 64.0, 70.0, 78.0, 84.0, 83.0, 80.0, 70.0, 55.0, 24.0])

        corrected = _apply_morning_shape_correction(final, baseline, hist, feat, capacity_mw=85.0)

        self.assertGreater(corrected[1], final[1])
        self.assertGreaterEqual(corrected[1], 50.0)
        self.assertLessEqual(corrected.sum(), final.sum() * 1.02)
        self.assertTrue(np.allclose(corrected[5:12], corrected[5:12].clip(max=85.0)))


class SingleAxisTrackerPostProcessingTests(TestCase):
    def setUp(self):
        user = User.objects.create_user(username="tracker", password="pass")
        org = Organization.objects.create(name="Tracker Org", owner=user)
        self.station = Station.objects.create(
            org=org,
            name="Tracker SES",
            mount_type=Station.MOUNT_SINGLE_AXIS_TRACKER,
            capacity_mw=1.0,
            capacity_ac_kw=1000,
            capacity_dc_kw=1100,
        )

    def test_tracker_profile_boosts_shoulders_flattens_midday_and_caps(self):
        tz = timezone.get_current_timezone()
        for day in range(1, 5):
            for hour in range(7, 18):
                plateau = 8 <= hour <= 15
                SolarRecord.objects.create(
                    station=self.station,
                    timestamp=timezone.datetime(2026, 5, day, hour, 0, tzinfo=tz),
                    history_scope=SolarRecord.HISTORY_SCOPE_MAIN,
                    irradiation=760.0 if plateau else 420.0,
                    air_temp=24.0,
                    power_kw=680.0 if plateau else 430.0,
                )

        feat = pd.DataFrame(
            {
                "ds": pd.to_datetime([
                    "2026-06-01 08:00:00",
                    "2026-06-01 12:00:00",
                    "2026-06-01 17:00:00",
                ]),
                "Irradiation": [650.0, 900.0, 500.0],
                "sun_elev_deg": [18.0, 65.0, 20.0],
            }
        )
        y = np.array([0.40, 0.80, 0.40])

        corrected, caps = _apply_single_axis_tracker_postprocessing(
            y,
            feat,
            self.station,
            capacity_mw=1.5,
        )

        self.assertGreater(corrected[0], 0.55)
        self.assertLess(corrected[1], y[1])
        self.assertGreater(corrected[2], y[2])
        self.assertLessEqual(float(corrected.max()), 1.0)
        self.assertEqual(caps["ac_cap_mw"], 1.0)
        self.assertLessEqual(caps["safe_cap_mw"], 1.0)
        self.assertGreaterEqual(caps["historical_profile_hours"], 8.0)
        self.assertGreaterEqual(caps["plateau_hours_applied"], 2.0)
        self.assertIn("tracker_profile_factor", feat.columns)
        self.assertIn("tracker_reference_mw", feat.columns)
        self.assertIn("tracker_station_calibration_factor", feat.columns)
        self.assertIn("tracker_calibrated_expected_mw", feat.columns)
        self.assertGreaterEqual(caps["station_calibration_hours"], 8.0)


class TrackerMiddayFloorTests(TestCase):
    def setUp(self):
        user = User.objects.create_user(username="shu-floor", password="pass")
        org = Organization.objects.create(name="Shu Org", owner=user)
        self.station = Station.objects.create(
            org=org,
            name="Tracker 100 MW",
            mount_type=Station.MOUNT_SINGLE_AXIS_TRACKER,
            capacity_mw=84.0,
            capacity_ac_kw=85000,
            capacity_dc_kw=100000,
        )

    def test_applies_midday_floor_for_high_irradiation(self):
        tz = timezone.get_current_timezone()
        for day in range(1, 5):
            for hour in range(9, 17):
                SolarRecord.objects.create(
                    station=self.station,
                    timestamp=timezone.datetime(2026, 5, day, hour, 0, tzinfo=tz),
                    history_scope=SolarRecord.HISTORY_SCOPE_MAIN,
                    irradiation=820.0,
                    air_temp=25.0,
                    power_kw=74000.0,
                )

        feat = pd.DataFrame(
            {
                "ds": pd.to_datetime(["2026-06-01 12:00:00", "2026-06-01 08:00:00"]),
                "Irradiation": [830.0, 830.0],
            }
        )
        y = np.array([55.0, 40.0], dtype=float)

        corrected = _apply_tracker_midday_expected_floor(y, feat, self.station, ac_cap_mw=85.0)

        self.assertGreaterEqual(corrected[0], 74.0)
        self.assertEqual(corrected[1], 40.0)

    def test_does_not_apply_for_fixed_mount(self):
        fixed_station = Station.objects.create(
            org=self.station.org,
            name="Fixed 50 MW",
            mount_type=Station.MOUNT_FIXED,
            capacity_mw=50.0,
            capacity_ac_kw=50000,
            capacity_dc_kw=60000,
        )
        feat = pd.DataFrame(
            {
                "ds": pd.to_datetime(["2026-06-01 12:00:00"]),
                "Irradiation": [900.0],
            }
        )
        y = np.array([30.0], dtype=float)
        corrected = _apply_tracker_midday_expected_floor(y, feat, fixed_station, ac_cap_mw=50.0)
        self.assertEqual(float(corrected[0]), 30.0)




class ForecastModeDiagnosticsTests(TestCase):
    def test_compare_forecast_modes_returns_hourly_values_and_logs_large_diffs(self):
        user = User.objects.create_user(username="diag", password="pass")
        org = Organization.objects.create(name="Diag Org", owner=user)
        station = Station.objects.create(
            org=org,
            name="Tracker 100 MW",
            mount_type=Station.MOUNT_SINGLE_AXIS_TRACKER,
            capacity_mw=84.0,
            capacity_ac_kw=85000,
            capacity_dc_kw=100000,
        )
        ts = timezone.datetime(2026, 5, 25, 12, 0, tzinfo=timezone.get_current_timezone())
        SolarForecast.objects.create(
            station=station,
            timestamp=ts,
            forecast_scope=SolarForecast.SCOPE_MAIN,
            irradiation_fc=700.0,
            pred_final_raw=60000.0,
            pred_final=61000.0,
        )
        SolarForecast.objects.create(
            station=station,
            timestamp=ts,
            forecast_scope=SolarForecast.SCOPE_TEST,
            irradiation_fc=900.0,
            pred_final_raw=74000.0,
            pred_final=75000.0,
        )

        with self.assertLogs("dashboard.services.forecast_diagnostics", level="WARNING") as logs:
            rows = compare_forecast_modes(station, date(2026, 5, 25), date(2026, 5, 31))

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].forecast_weather_irr, 700.0)
        self.assertEqual(rows[0].postfact_weather_irr, 900.0)
        self.assertAlmostEqual(rows[0].forecast_raw_mw, 60.0)
        self.assertAlmostEqual(rows[0].postfact_final_mw, 75.0)
        self.assertIn("[FORECAST_COMPARE]", logs.output[0])

class StationAutoHistoryStandardFileTests(TestCase):
    def test_collect_share_history_uses_standard_history_columns_from_csv(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            pd.DataFrame(
                [
                    {
                        "ds": "2026-02-17 10:10:00",
                        "Irradiation": 500.456,
                        "Air_Temp": 20.222,
                        "PV_Temp": 25.111,
                        "Power_KW": 120.987,
                    }
                ]
            ).to_csv(folder / "history_upload.csv", index=False)

            out = collect_share_history_dataframe(folder)

        self.assertEqual(len(out), 1)
        self.assertEqual(str(out.iloc[0]["ds"]), "2026-02-17 10:00:00")
        self.assertAlmostEqual(float(out.iloc[0]["irradiation"]), 500.456, places=3)
        self.assertAlmostEqual(float(out.iloc[0]["air_temp"]), 20.222, places=3)
        self.assertAlmostEqual(float(out.iloc[0]["pv_temp"]), 25.111, places=3)
        self.assertAlmostEqual(float(out.iloc[0]["power_kw"]), 120.99, places=2)

    def test_collect_share_history_standard_file_supports_timestamp_alias(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            pd.DataFrame(
                [
                    {
                        "timestamp": "2026-02-17 11:40:00",
                        "Irradiation": 410,
                        "Air_Temp": 18,
                        "PV_Temp": 22,
                        "Power_KW": 0,
                    },
                    {
                        "timestamp": "2026-02-17 12:05:00",
                        "Irradiation": 430,
                        "Air_Temp": 19,
                        "PV_Temp": 23,
                        "Power_KW": 140,
                    },
                ]
            ).to_csv(folder / "history_upload.csv", index=False)

            out = collect_share_history_dataframe(folder)

        self.assertEqual(len(out), 1)
        self.assertEqual(str(out.iloc[0]["ds"]), "2026-02-17 12:00:00")
        self.assertAlmostEqual(float(out.iloc[0]["power_kw"]), 140.0)


class ForecastEngineXgbLowConfidenceTests(TestCase):
    def test_marks_xgb_low_when_it_is_much_lower_than_np_and_heuristic(self):
        y_xgb = [5.3, 5.2, 5.1]
        y_heur = [213.8, 215.7, 217.3]
        y_np = [258.2, 256.1, 258.8]

        low = _xgb_is_systematically_low(y_xgb, y_heur, y_np, np_ok=True, capacity_mw=50.0)

        self.assertTrue(low)

    def test_does_not_mark_low_when_np_is_unavailable(self):
        y_xgb = [5.3, 5.2, 5.1]
        y_heur = [213.8, 215.7, 217.3]
        y_np = [0.0, 0.0, 0.0]

        low = _xgb_is_systematically_low(y_xgb, y_heur, y_np, np_ok=False, capacity_mw=50.0)

        self.assertFalse(low)



    def test_tracker_training_features_make_xgb_target_use_tracker_expected_curve(self):
        station = SimpleNamespace(
            pk=11,
            mount_type=Station.MOUNT_SINGLE_AXIS_TRACKER,
        )
        df = pd.DataFrame(
            {
                "ds": pd.to_datetime([
                    "2026-05-01 08:00:00",
                    "2026-05-01 12:00:00",
                    "2026-05-01 16:00:00",
                    "2026-05-02 08:00:00",
                    "2026-05-02 12:00:00",
                    "2026-05-02 16:00:00",
                ]),
                "hour": [8, 12, 16, 8, 12, 16],
                "Irradiation": [700.0, 900.0, 650.0, 720.0, 880.0, 640.0],
                "Air_Temp": [22.0] * 6,
                "sun_elev_deg": [25.0, 65.0, 24.0, 25.0, 65.0, 24.0],
                "y_expected": [50.0] * 6,
                "y": [60.0, 58.0, 55.0, 61.0, 57.0, 54.0],
            }
        )

        out = add_tracker_training_features(df, station, cap_mw=85.0)
        prepared = _prepare_xgb_training_frame(out, cap_mw=85.0)

        self.assertIn("tracker_profile_factor", out.columns)
        self.assertIn("tracker_calibrated_expected_mw", out.columns)
        first_expected = float(out.loc[prepared.index[0], "tracker_calibrated_expected_mw"])
        self.assertAlmostEqual(
            float(prepared.iloc[0]["y_over_expected"]),
            min(float(out.loc[prepared.index[0], "y"]) / max(first_expected, 4.25), 1.6),
        )


class ForecastEngineIndexRegressionTests(TestCase):
    def setUp(self):
        user = User.objects.create_user(username="tester", password="pass")
        org = Organization.objects.create(name="Org", owner=user)
        self.station = Station.objects.create(
            org=org,
            name="SES 8.8MW",
            capacity_mw=8.8,
            capacity_ac_kw=8800,
            capacity_dc_kw=9000,
            latitude=None,
            longitude=None,
        )

    @patch("dashboard.services.forecast_engine._target_offsets_for_weekday_calendar", return_value=[2])
    def test_run_forecast_handles_sparse_index_after_target_date_filter(self, _offsets):
        """
        Regression: после фильтрации target_dates DataFrame feat может иметь индекс,
        начинающийся не с 0. Сохранение прогноза должно работать без IndexError.
        """
        result = run_forecast_for_station(
            station_id=self.station.pk,
            days=1,
            use_models=False,
            horizon_mode="weekday_calendar",
        )

        self.assertTrue(result["ok"])
        self.assertGreater(result["count"], 0)


class ForecastEngineHourlyWeatherFallbackTests(TestCase):
    def setUp(self):
        user = User.objects.create_user(username="hourly-fallback", password="pass")
        org = Organization.objects.create(name="Hourly Fallback Org", owner=user)
        self.station = Station.objects.create(
            org=org,
            name="Hourly Fallback Station",
            capacity_mw=1.0,
            capacity_ac_kw=1000,
            capacity_dc_kw=1100,
            latitude=50.0,
            longitude=30.0,
        )

    @patch("dashboard.services.forecast_engine.fetch_open_meteo_hourly")
    @patch("dashboard.services.forecast_engine.timezone.now")
    def test_run_forecast_uses_hourly_weather_profile_when_target_date_has_no_direct_weather(self, now_mock, meteo_mock):
        now = timezone.datetime(2026, 2, 25, 15, 0, tzinfo=timezone.get_current_timezone())
        now_mock.return_value = now

        weather_df = pd.DataFrame(
            {
                "ds": [
                    timezone.datetime(2026, 2, 26, 9, 0, tzinfo=timezone.get_current_timezone()),
                    timezone.datetime(2026, 2, 26, 10, 0, tzinfo=timezone.get_current_timezone()),
                ],
                "irradiation": [400.0, 500.0],
                "air_temp": [8.0, 10.0],
                "wind_speed": [2.0, 2.5],
                "cloudcover": [30.0, 20.0],
                "humidity": [55.0, 50.0],
                "precip": [0.0, 0.0],
                "snowfall": [0.0, 0.0],
                "snowdepth": [0.0, 0.0],
                "weather_code": [1, 1],
            }
        )
        meteo_mock.return_value = WeatherFetchResult(ok=True, source="open_meteo", df=weather_df)

        result = run_forecast_for_station(
            station_id=self.station.pk,
            days=1,
            providers=["open_meteo"],
            use_models=False,
            target_dates=[date(2026, 2, 24)],
        )

        self.assertTrue(result["ok"])
        forecasts = SolarForecast.objects.filter(station=self.station)
        self.assertTrue(forecasts.exists())
        self.assertGreater(forecasts.aggregate(Max("pred_heur"))["pred_heur__max"], 0)


class ForecastEngineHistoryBackfillFallbackTests(TestCase):
    def setUp(self):
        user = User.objects.create_user(username="history-fallback", password="pass")
        org = Organization.objects.create(name="History Fallback Org", owner=user)
        self.station = Station.objects.create(
            org=org,
            name="Fallback Station",
            capacity_mw=1.0,
            capacity_ac_kw=1000,
            capacity_dc_kw=1100,
            latitude=None,
            longitude=None,
        )

    @patch("dashboard.services.forecast_engine.fetch_open_meteo_hourly")
    @patch("dashboard.services.forecast_engine.timezone.now")
    def test_run_forecast_keeps_provider_weather_when_history_backfill_is_empty(self, now_mock, meteo_mock):
        self.station.latitude = 50.0
        self.station.longitude = 30.0
        self.station.save(update_fields=["latitude", "longitude"])

        now = timezone.datetime(2026, 2, 25, 15, 0, tzinfo=timezone.get_current_timezone())
        now_mock.return_value = now

        weather_df = pd.DataFrame(
            {
                "ds": [
                    timezone.datetime(2026, 2, 26, 11, 0, tzinfo=timezone.get_current_timezone()),
                    timezone.datetime(2026, 2, 26, 12, 0, tzinfo=timezone.get_current_timezone()),
                ],
                "irradiation": [450.0, 600.0],
                "air_temp": [9.0, 11.0],
                "wind_speed": [2.0, 2.3],
                "cloudcover": [20.0, 10.0],
                "humidity": [55.0, 50.0],
                "precip": [0.0, 0.0],
                "snowfall": [0.0, 0.0],
                "snowdepth": [0.0, 0.0],
                "weather_code": [1, 1],
            }
        )
        meteo_mock.return_value = WeatherFetchResult(ok=True, source="open_meteo", df=weather_df)

        result = run_forecast_for_station(
            station_id=self.station.pk,
            days=1,
            providers=["open_meteo"],
            use_models=False,
            forecast_scope="test",
            target_dates=[date(2026, 2, 24)],
        )

        self.assertTrue(result["ok"])
        self.assertEqual(result["weather_source"], "open_meteo")
        forecasts = SolarForecast.objects.filter(station=self.station)
        self.assertTrue(forecasts.exists())
        self.assertGreater(forecasts.aggregate(Max("pred_heur"))["pred_heur__max"], 0)

    @patch("dashboard.services.forecast_engine.fetch_visual_crossing_hourly")
    @patch("dashboard.services.forecast_engine.timezone.now")
    def test_run_forecast_tolerates_provider_weather_without_optional_columns(self, now_mock, crossing_mock):
        self.station.latitude = 50.0
        self.station.longitude = 30.0
        self.station.save(update_fields=["latitude", "longitude"])

        now = timezone.datetime(2026, 2, 25, 15, 0, tzinfo=timezone.get_current_timezone())
        now_mock.return_value = now

        weather_df = pd.DataFrame(
            {
                "ds": [
                    timezone.datetime(2026, 2, 26, 11, 0, tzinfo=timezone.get_current_timezone()),
                    timezone.datetime(2026, 2, 26, 12, 0, tzinfo=timezone.get_current_timezone()),
                ],
                "irradiation": [450.0, 600.0],
                "air_temp": [9.0, 11.0],
                "wind_speed": [2.0, 2.3],
                "cloudcover": [20.0, 10.0],
                "humidity": [55.0, 50.0],
                "precip": [0.0, 0.0],
            }
        )
        crossing_mock.return_value = WeatherFetchResult(ok=True, source="visual_crossing", df=weather_df)

        result = run_forecast_for_station(
            station_id=self.station.pk,
            days=1,
            providers=["visual_crossing"],
            use_models=False,
            forecast_scope="test",
            target_dates=[date(2026, 2, 24)],
        )

        self.assertTrue(result["ok"])
        self.assertEqual(result["weather_source"], "visual_crossing")
        forecasts = SolarForecast.objects.filter(station=self.station)
        self.assertTrue(forecasts.exists())
        self.assertGreater(forecasts.aggregate(Max("pred_heur"))["pred_heur__max"], 0)

    @patch("dashboard.services.forecast_engine.timezone.now")
    def test_run_forecast_uses_main_history_when_test_scope_missing(self, now_mock):
        now = timezone.datetime(2026, 2, 25, 15, 0, tzinfo=timezone.get_current_timezone())
        now_mock.return_value = now
        SolarRecord.objects.create(
            station=self.station,
            timestamp=timezone.datetime(2026, 2, 25, 12, 0, tzinfo=timezone.get_current_timezone()),
            history_scope=SolarRecord.HISTORY_SCOPE_MAIN,
            irradiation=620.0,
            air_temp=11.0,
        )

        result = run_forecast_for_station(
            station_id=self.station.pk,
            days=1,
            use_models=False,
            forecast_scope="test",
            target_dates=[date(2026, 2, 25)],
        )

        self.assertTrue(result["ok"])
        self.assertEqual(result["weather_source"], "history_backfill")
        self.assertGreater(result["count"], 0)




class ForecastEngineManualSnowFactorIncreaseTests(TestCase):
    def setUp(self):
        user = User.objects.create_user(username="manual-snow-up", password="pass")
        org = Organization.objects.create(name="Manual Snow Org", owner=user)
        self.station = Station.objects.create(
            org=org,
            name="Manual Snow Station",
            capacity_mw=1.0,
            capacity_ac_kw=1000,
            capacity_dc_kw=1100,
            latitude=None,
            longitude=None,
        )

    @patch("dashboard.services.forecast_engine.timezone.now")
    def test_manual_snow_factor_can_increase_forecast_but_not_above_station_capacity(self, now_mock):
        now = timezone.datetime(2026, 2, 25, 15, 0, tzinfo=timezone.get_current_timezone())
        now_mock.return_value = now
        SolarRecord.objects.create(
            station=self.station,
            timestamp=timezone.datetime(2026, 2, 25, 12, 0, tzinfo=timezone.get_current_timezone()),
            history_scope=SolarRecord.HISTORY_SCOPE_MAIN,
            irradiation=620.0,
            air_temp=11.0,
        )

        baseline = run_forecast_for_station(
            station_id=self.station.pk,
            days=1,
            use_models=False,
            forecast_scope="test",
            target_dates=[date(2026, 2, 25)],
        )
        self.assertTrue(baseline["ok"])
        baseline_max = SolarForecast.objects.filter(station=self.station).aggregate(Max("pred_final"))["pred_final__max"]

        increased = run_forecast_for_station(
            station_id=self.station.pk,
            days=1,
            use_models=False,
            forecast_scope="test",
            target_dates=[date(2026, 2, 25)],
            manual_snow_enable=True,
            manual_snow_factor=1.15,
        )
        self.assertTrue(increased["ok"])

        boosted_max = SolarForecast.objects.filter(station=self.station).aggregate(Max("pred_final"))["pred_final__max"]

        self.assertGreaterEqual(boosted_max, baseline_max)
        self.assertLessEqual(boosted_max, self.station.capacity_mw * 1000)

    @patch("dashboard.services.forecast_engine.timezone.now")
    def test_manual_snow_factor_is_clipped_to_safe_upper_bound(self, now_mock):
        now = timezone.datetime(2026, 2, 25, 15, 0, tzinfo=timezone.get_current_timezone())
        now_mock.return_value = now
        SolarRecord.objects.create(
            station=self.station,
            timestamp=timezone.datetime(2026, 2, 25, 12, 0, tzinfo=timezone.get_current_timezone()),
            history_scope=SolarRecord.HISTORY_SCOPE_MAIN,
            irradiation=620.0,
            air_temp=11.0,
        )

        result = run_forecast_for_station(
            station_id=self.station.pk,
            days=1,
            use_models=False,
            forecast_scope="test",
            target_dates=[date(2026, 2, 25)],
            manual_snow_enable=True,
            manual_snow_factor=9.0,
        )

        self.assertTrue(result["ok"])
        applied_factor = SolarForecast.objects.filter(station=self.station).aggregate(Max("winter_factor_applied"))["winter_factor_applied__max"]
        self.assertEqual(applied_factor, 1.5)


class ForecastEngineGlobalBiasTests(TestCase):
    def setUp(self):
        user = User.objects.create_user(username="global-bias", password="pass")
        org = Organization.objects.create(name="Global Bias Org", owner=user)
        self.station = Station.objects.create(
            org=org,
            name="Bias Station",
            capacity_mw=1.0,
            capacity_ac_kw=1000,
            capacity_dc_kw=1100,
            latitude=None,
            longitude=None,
        )
        SolarRecord.objects.create(
            station=self.station,
            timestamp=timezone.datetime(2026, 2, 25, 12, 0, tzinfo=timezone.get_current_timezone()),
            history_scope=SolarRecord.HISTORY_SCOPE_MAIN,
            irradiation=620.0,
            air_temp=11.0,
        )

    @patch("dashboard.services.forecast_engine.timezone.now")
    def test_global_bias_increases_final_forecast(self, now_mock):
        now = timezone.datetime(2026, 2, 25, 15, 0, tzinfo=timezone.get_current_timezone())
        now_mock.return_value = now

        baseline = run_forecast_for_station(
            station_id=self.station.pk,
            days=1,
            use_models=False,
            forecast_scope="test",
            target_dates=[date(2026, 2, 25)],
        )
        self.assertTrue(baseline["ok"])
        baseline_max = SolarForecast.objects.filter(station=self.station).aggregate(Max("pred_final"))["pred_final__max"]

        with override_settings(FORECAST_GLOBAL_BIAS=1.10):
            boosted = run_forecast_for_station(
                station_id=self.station.pk,
                days=1,
                use_models=False,
                forecast_scope="test",
                target_dates=[date(2026, 2, 25)],
            )
        self.assertTrue(boosted["ok"])
        boosted_max = SolarForecast.objects.filter(station=self.station).aggregate(Max("pred_final"))["pred_final__max"]

        self.assertGreater(boosted_max, baseline_max)

    @patch("dashboard.services.forecast_engine.timezone.now")
    def test_global_bias_is_clipped_by_capacity_and_safe_limit(self, now_mock):
        now = timezone.datetime(2026, 2, 25, 15, 0, tzinfo=timezone.get_current_timezone())
        now_mock.return_value = now

        with override_settings(FORECAST_GLOBAL_BIAS=9.0):
            result = run_forecast_for_station(
                station_id=self.station.pk,
                days=1,
                use_models=False,
                forecast_scope="test",
                target_dates=[date(2026, 2, 25)],
            )

        self.assertTrue(result["ok"])
        max_pred_final = SolarForecast.objects.filter(station=self.station).aggregate(Max("pred_final"))["pred_final__max"]
        self.assertLessEqual(max_pred_final, self.station.capacity_mw * 1000)


class ForecastEngineWeekdayCalendarTests(TestCase):
    def test_friday_offsets_skip_saturday_and_cover_sun_mon_tue(self):
        friday = timezone.datetime(2026, 2, 13, 9, 0)

        offsets = _target_offsets_for_weekday_calendar(friday)

        self.assertEqual(offsets, [2, 3, 4])

    def test_weekday_mode_report_days_do_not_depend_on_requested_days(self):
        user = User.objects.create_user(username="daysmode", password="pass")
        org = Organization.objects.create(name="Days Mode Org", owner=user)
        station = Station.objects.create(
            org=org,
            name="Days Mode Station",
            capacity_mw=1.0,
            capacity_ac_kw=1000,
            capacity_dc_kw=1100,
            latitude=None,
            longitude=None,
        )

        with patch("dashboard.services.forecast_engine._target_offsets_for_weekday_calendar", return_value=[2, 3, 4]):
            result_days_1 = run_forecast_for_station(
                station_id=station.pk,
                days=1,
                use_models=False,
                horizon_mode="weekday_calendar",
            )
            result_days_7 = run_forecast_for_station(
                station_id=station.pk,
                days=7,
                use_models=False,
                horizon_mode="weekday_calendar",
            )

        self.assertTrue(result_days_1["ok"])
        self.assertTrue(result_days_7["ok"])
        self.assertEqual(result_days_1["days"], 3)
        self.assertEqual(result_days_7["days"], 3)


class ForecastSchedulerForceRunTests(TestCase):
    def setUp(self):
        user = User.objects.create_user(username="scheduler", password="pass")
        org = Organization.objects.create(name="Scheduler Org", owner=user)
        self.station = Station.objects.create(
            org=org,
            name="Scheduler Station",
            capacity_mw=1.0,
            capacity_ac_kw=1000,
            capacity_dc_kw=1100,
        )
        self.schedule = ForecastSchedule.objects.create(
            station=self.station,
            enabled=True,
            days=1,
            run_time=timezone.datetime.strptime("23:59", "%H:%M").time(),
            emails="test@example.com",
        )


    @patch("dashboard.services.forecast_scheduler.send_report_email", return_value=True)
    @patch("dashboard.services.forecast_scheduler.build_forecast_report", return_value=object())
    @patch(
        "dashboard.services.forecast_scheduler.run_forecast_for_station",
        return_value={"ok": True, "weather_source": "stub", "target_dates": ["2026-02-13"]},
    )
    def test_scheduler_saves_last_email_delivery_info(self, _run, _build, _send):
        now = timezone.now().replace(hour=23, minute=59, second=0, microsecond=0)

        count = run_scheduled_forecasts(now=now, force=True)
        self.schedule.refresh_from_db()

        self.assertEqual(count, 1)
        self.assertIsNotNone(self.schedule.last_email_sent_at)
        self.assertEqual(str(self.schedule.last_email_forecast_date), "2026-02-13")
        self.assertIn("Прогноз за 13.02.2026 отправлен", self.schedule.last_email_status)

    @patch("dashboard.services.forecast_scheduler.send_report_email")
    @patch("dashboard.services.forecast_scheduler.build_forecast_report", return_value=object())
    @patch(
        "dashboard.services.forecast_scheduler.run_forecast_for_station",
        return_value={"ok": True, "weather_source": "stub"},
    )
    def test_scheduler_uses_main_scope(self, run_mock, build_mock, _send):
        now = timezone.now().replace(hour=23, minute=59, second=0, microsecond=0)

        count = run_scheduled_forecasts(now=now, force=True)

        self.assertEqual(count, 1)
        self.assertEqual(run_mock.call_args.kwargs["forecast_scope"], "main")
        self.assertEqual(build_mock.call_args.kwargs["forecast_scope"], "main")

    @patch("dashboard.services.forecast_scheduler.send_report_email")
    @patch("dashboard.services.forecast_scheduler.build_forecast_report", return_value=object())
    @patch(
        "dashboard.services.forecast_scheduler.run_forecast_for_station",
        return_value={"ok": True, "weather_source": "stub"},
    )
    def test_force_run_ignores_time_and_last_run_limit(self, _run, _build, _send):
        now = timezone.now().replace(hour=9, minute=0, second=0, microsecond=0)

        first = run_scheduled_forecasts(now=now, force=True)
        second = run_scheduled_forecasts(now=now, force=True)

        self.assertEqual(first, 1)
        self.assertEqual(second, 1)


    @patch("dashboard.services.forecast_scheduler.send_report_email")
    @patch("dashboard.services.forecast_scheduler.build_forecast_report", return_value=object())
    @patch(
        "dashboard.services.forecast_scheduler.run_forecast_for_station",
        return_value={"ok": True, "weather_source": "stub", "days": 3, "target_dates": ["2026-02-15", "2026-02-16", "2026-02-17"]},
    )
    def test_report_uses_effective_days_from_engine_result(self, _run, build_mock, send_mock):
        now = timezone.now().replace(hour=23, minute=59, second=0, microsecond=0)

        count = run_scheduled_forecasts(now=now, force=True)

        self.assertEqual(count, 1)
        build_mock.assert_called_once()
        self.assertEqual(build_mock.call_args.kwargs["days"], 3)
        self.assertEqual(
            build_mock.call_args.kwargs["target_dates"],
            ["2026-02-15", "2026-02-16", "2026-02-17"],
        )
        send_mock.assert_called_once()
        self.assertEqual(send_mock.call_args.args[3], 3)


    @patch("dashboard.services.forecast_scheduler.send_report_email")
    @patch("dashboard.services.forecast_scheduler.build_forecast_report", return_value=object())
    @patch(
        "dashboard.services.forecast_scheduler.run_forecast_for_station",
        return_value={"ok": True, "weather_source": "stub"},
    )
    def test_first_run_respects_run_time_even_when_start_at_passed(self, run_mock, _build, _send):
        today = timezone.localtime(timezone.now()).date()
        self.schedule.start_at = timezone.make_aware(timezone.datetime.combine(today, timezone.datetime.strptime("00:01", "%H:%M").time()))
        self.schedule.run_time = timezone.datetime.strptime("14:05", "%H:%M").time()
        self.schedule.last_run_at = None
        self.schedule.save(update_fields=["start_at", "run_time", "last_run_at"])

        now_before = timezone.make_aware(timezone.datetime.combine(today, timezone.datetime.strptime("14:00", "%H:%M").time()))
        now_due = timezone.make_aware(timezone.datetime.combine(today, timezone.datetime.strptime("14:05", "%H:%M").time()))

        first = run_scheduled_forecasts(now=now_before, force=False)
        second = run_scheduled_forecasts(now=now_due, force=False)

        self.assertEqual(first, 0)
        self.assertEqual(second, 1)
        self.assertEqual(run_mock.call_count, 1)

    @patch("dashboard.services.forecast_scheduler.send_report_email")
    @patch("dashboard.services.forecast_scheduler.build_forecast_report", return_value=object())
    @patch(
        "dashboard.services.forecast_scheduler.run_forecast_for_station",
        return_value={"ok": True, "weather_source": "stub"},
    )
    def test_non_force_respects_daily_limit(self, _run, _build, _send):
        now = timezone.now().replace(hour=23, minute=59, second=0, microsecond=0)

        first = run_scheduled_forecasts(now=now, force=False)
        second = run_scheduled_forecasts(now=now, force=False)

        self.assertEqual(first, 1)
        self.assertEqual(second, 0)

    @patch("dashboard.services.forecast_scheduler.send_report_email")
    @patch("dashboard.services.forecast_scheduler.build_forecast_report", return_value=object())
    @patch(
        "dashboard.services.forecast_scheduler.run_forecast_for_station",
        return_value={"ok": True, "weather_source": "stub", "days": 3},
    )
    def test_schedule_open_meteo_only_disables_models(self, run_mock, _build, _send):
        self.schedule.providers = "visual_crossing,open_meteo_only"
        self.schedule.save(update_fields=["providers"])

        now = timezone.now().replace(hour=23, minute=59, second=0, microsecond=0)
        count = run_scheduled_forecasts(now=now, force=True)

        self.assertEqual(count, 1)
        self.assertEqual(run_mock.call_args.kwargs["providers"], ["visual_crossing"])
        self.assertFalse(run_mock.call_args.kwargs["use_models"])

    @patch("dashboard.services.forecast_scheduler.send_report_email")
    @patch("dashboard.services.forecast_scheduler.build_forecast_report", return_value=object())
    @patch(
        "dashboard.services.forecast_scheduler.run_forecast_for_station",
        return_value={"ok": True, "weather_source": "stub", "days": 3},
    )
    def test_schedule_visual_crossing_only_disables_models(self, run_mock, _build, _send):
        self.schedule.providers = "visual_crossing_only"
        self.schedule.save(update_fields=["providers"])

        now = timezone.now().replace(hour=23, minute=59, second=0, microsecond=0)
        count = run_scheduled_forecasts(now=now, force=True)

        self.assertEqual(count, 1)
        self.assertEqual(run_mock.call_args.kwargs["providers"], ["visual_crossing"])
        self.assertFalse(run_mock.call_args.kwargs["use_models"])


    @patch("dashboard.services.forecast_scheduler.send_report_email")
    @patch("dashboard.services.forecast_scheduler.build_forecast_report", return_value=object())
    @patch(
        "dashboard.services.forecast_scheduler.run_forecast_for_station",
        return_value={"ok": False, "error": "provider unavailable"},
    )
    def test_failed_run_does_not_mark_schedule_as_executed(self, run_mock, _build, _send):
        now = timezone.now().replace(hour=23, minute=59, second=0, microsecond=0)

        first = run_scheduled_forecasts(now=now, force=False)
        self.schedule.refresh_from_db()
        second = run_scheduled_forecasts(now=now, force=False)

        self.assertEqual(first, 0)
        self.assertEqual(second, 0)
        self.assertIsNone(self.schedule.last_run_at)
        self.assertEqual(run_mock.call_count, 2)


    @override_settings(FORECAST_EMAIL_MAX_ATTEMPTS=2, FORECAST_EMAIL_RETRY_DELAY_SECONDS=0)
    @patch("dashboard.services.forecast_scheduler.send_report_email", return_value=False)
    @patch("dashboard.services.forecast_scheduler.build_forecast_report", return_value=object())
    @patch(
        "dashboard.services.forecast_scheduler.run_forecast_for_station",
        return_value={"ok": True, "weather_source": "stub"},
    )
    def test_email_failure_does_not_mark_schedule_as_executed(self, run_mock, _build, send_mock):
        now = timezone.now().replace(hour=23, minute=59, second=0, microsecond=0)

        first = run_scheduled_forecasts(now=now, force=False)
        self.schedule.refresh_from_db()
        second = run_scheduled_forecasts(now=now, force=False)

        self.assertEqual(first, 0)
        self.assertEqual(second, 0)
        self.assertIsNone(self.schedule.last_run_at)
        self.assertEqual(run_mock.call_count, 2)
        self.assertEqual(send_mock.call_count, 4)



    @override_settings(FORECAST_EMAIL_MAX_ATTEMPTS=3, FORECAST_EMAIL_RETRY_DELAY_SECONDS=0)
    @patch("dashboard.services.forecast_scheduler.send_report_email", side_effect=[False, True])
    @patch("dashboard.services.forecast_scheduler.build_forecast_report", return_value=object())
    @patch(
        "dashboard.services.forecast_scheduler.run_forecast_for_station",
        return_value={"ok": True, "weather_source": "stub"},
    )
    def test_email_retry_success_marks_schedule_as_executed(self, run_mock, _build, send_mock):
        now = timezone.now().replace(hour=23, minute=59, second=0, microsecond=0)

        count = run_scheduled_forecasts(now=now, force=False)
        self.schedule.refresh_from_db()

        self.assertEqual(count, 1)
        self.assertIsNotNone(self.schedule.last_run_at)
        self.assertEqual(run_mock.call_count, 1)
        self.assertEqual(send_mock.call_count, 2)

    @patch("dashboard.services.forecast_scheduler.send_report_email")
    @patch("dashboard.services.forecast_scheduler.build_forecast_report", return_value=object())
    @patch(
        "dashboard.services.forecast_scheduler.run_forecast_for_station",
        return_value={"ok": True, "weather_source": "stub"},
    )
    def test_schedule_without_emails_marks_run_without_sending(self, run_mock, _build, send_mock):
        self.schedule.emails = ""
        self.schedule.save(update_fields=["emails"])

        now = timezone.now().replace(hour=23, minute=59, second=0, microsecond=0)
        count = run_scheduled_forecasts(now=now, force=False)
        self.schedule.refresh_from_db()

        self.assertEqual(count, 1)
        self.assertIsNotNone(self.schedule.last_run_at)
        self.assertEqual(run_mock.call_count, 1)
        send_mock.assert_not_called()


class StationForecastRunTargetDatesTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="run-target-dates", password="pass")
        self.org = Organization.objects.create(name="Run Target Dates Org", owner=self.user)
        OrganizationMember.objects.create(
            organization=self.org,
            user=self.user,
            role=OrganizationMember.ROLE_OWNER,
        )
        self.station = Station.objects.create(org=self.org, name="Target Dates Station", capacity_mw=1.2)
        self.client.login(username="run-target-dates", password="pass")

    @patch("dashboard.views.send_report_email", return_value=False)
    @patch("dashboard.views.build_forecast_report")
    @patch("dashboard.views.run_forecast_for_station", return_value={"ok": True, "weather_source": "stub", "days": 2, "target_dates": ["2026-02-22", "2026-02-23"]})
    def test_station_forecast_run_passes_target_dates(self, run_mock, _build, _send):
        response = self.client.get(
            f"/dashboard/station/{self.station.pk}/forecast/run/",
            {
                "days": "2",
                "scope": "main",
                "target_dates": "2026-02-22, 2026-02-23",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            run_mock.call_args.kwargs["target_dates"],
            [date(2026, 2, 22), date(2026, 2, 23)],
        )

    @patch("dashboard.views.send_report_email", return_value=False)
    @patch("dashboard.views.build_forecast_report")
    @patch("dashboard.views.run_forecast_for_station", return_value={"ok": True, "weather_source": "stub", "days": 2})
    def test_station_forecast_run_open_meteo_only_keeps_explicit_providers(self, run_mock, _build, _send):
        response = self.client.get(
            f"/dashboard/station/{self.station.pk}/forecast/run/",
            {
                "days": "2",
                "scope": "main",
                "providers": ["visual_crossing"],
                "open_meteo_only": "1",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(run_mock.call_args.kwargs["providers"], ["visual_crossing"])
        self.assertFalse(run_mock.call_args.kwargs["use_models"])

    @patch("dashboard.views.send_report_email", return_value=False)
    @patch("dashboard.views.build_forecast_report")
    @patch("dashboard.views.run_forecast_for_station", return_value={"ok": True, "weather_source": "stub", "days": 2})
    def test_station_forecast_run_open_meteo_only_defaults_to_open_meteo_provider(self, run_mock, _build, _send):
        response = self.client.get(
            f"/dashboard/station/{self.station.pk}/forecast/run/",
            {
                "days": "2",
                "scope": "main",
                "open_meteo_only": "1",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(run_mock.call_args.kwargs["providers"], ["open_meteo"])
        self.assertFalse(run_mock.call_args.kwargs["use_models"])

    @patch("dashboard.views.send_report_email", return_value=False)
    @patch("dashboard.views.build_forecast_report")
    @patch("dashboard.views.run_forecast_for_station", return_value={"ok": True, "weather_source": "stub", "days": 2})
    def test_station_forecast_run_visual_crossing_only_defaults_to_visual_crossing_provider(self, run_mock, _build, _send):
        response = self.client.get(
            f"/dashboard/station/{self.station.pk}/forecast/run/",
            {
                "days": "2",
                "scope": "main",
                "visual_crossing_only": "1",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(run_mock.call_args.kwargs["providers"], ["visual_crossing"])
        self.assertFalse(run_mock.call_args.kwargs["use_models"])

    @patch("dashboard.views.send_report_email", return_value=False)
    @patch("dashboard.views.build_forecast_report")
    @patch("dashboard.views.run_forecast_for_station")
    def test_station_forecast_run_retries_without_models_on_failure(self, run_mock, _build, _send):
        run_mock.side_effect = [
            RuntimeError("model crash"),
            {"ok": True, "weather_source": "stub", "days": 2},
        ]

        response = self.client.get(
            f"/dashboard/station/{self.station.pk}/forecast/run/",
            {
                "days": "2",
                "scope": "main",
                "providers": ["visual_crossing"],
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(run_mock.call_count, 2)
        self.assertTrue(run_mock.call_args_list[0].kwargs["use_models"])
        self.assertFalse(run_mock.call_args_list[1].kwargs["use_models"])

    @patch("dashboard.views.send_report_email", return_value=False)
    @patch("dashboard.views.build_forecast_report")
    @patch("dashboard.views.run_forecast_for_station", return_value={"ok": True, "weather_source": "stub", "days": 1, "target_dates": ["2026-02-22"]})
    def test_station_forecast_run_target_dates_override_days(self, run_mock, build_mock, _send):
        response = self.client.get(
            f"/dashboard/station/{self.station.pk}/forecast/run/",
            {
                "days": "7",
                "scope": "test",
                "target_dates": "2026-02-22, 2026-02-22",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(run_mock.call_args.kwargs["days"], 1)
        self.assertEqual(run_mock.call_args.kwargs["target_dates"], [date(2026, 2, 22)])
        self.assertEqual(build_mock.call_args.kwargs["days"], 1)


class ForecastSchedulerProviderNormalizationTests(TestCase):
    def test_open_meteo_only_provider_marker_forces_open_meteo_and_heuristic(self):
        providers, open_meteo_only = _normalize_schedule_providers("visual_crossing,open_meteo_only")

        self.assertEqual(providers, ["visual_crossing"])
        self.assertTrue(open_meteo_only)

    def test_open_meteo_only_without_explicit_provider_falls_back_to_open_meteo(self):
        providers, open_meteo_only = _normalize_schedule_providers("open_meteo_only")

        self.assertEqual(providers, ["open_meteo"])
        self.assertTrue(open_meteo_only)

    def test_visual_crossing_only_without_explicit_provider_falls_back_to_visual_crossing(self):
        providers, open_meteo_only = _normalize_schedule_providers("visual_crossing_only")

        self.assertEqual(providers, ["visual_crossing"])
        self.assertTrue(open_meteo_only)


class ForecastSchedulerTickViewTests(TestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.user = User.objects.create_user(username="viewer", password="pass")

    @patch("dashboard.views.run_auto_history_updates", return_value=5)
    @patch("dashboard.views.run_scheduled_forecasts", return_value=3)
    def test_scheduler_tick_parses_force_true(self, run_mock, history_mock):
        request = self.factory.get("/dashboard/stations/1/forecast/scheduler-tick/?force=1")
        request.user = self.user

        response = station_forecast_scheduler_tick(request)

        self.assertEqual(response.status_code, 200)
        history_mock.assert_called_once_with()
        run_mock.assert_called_once_with(force=True)
        self.assertIn(b'"force": true', response.content)
        self.assertIn(b'"auto_history_rows": 5', response.content)
        self.assertIn(b'"forecast_count": 3', response.content)

    @patch("dashboard.views.run_auto_history_updates", return_value=0)
    @patch("dashboard.views.run_scheduled_forecasts", return_value=0)
    def test_scheduler_tick_defaults_force_false(self, run_mock, history_mock):
        request = self.factory.get("/dashboard/stations/1/forecast/scheduler-tick/")
        request.user = self.user

        response = station_forecast_scheduler_tick(request)

        self.assertEqual(response.status_code, 200)
        history_mock.assert_called_once_with()
        run_mock.assert_called_once_with(force=False)
        self.assertIn(b'"force": false', response.content)
        self.assertIn(b'"auto_history_rows": 0', response.content)


class StationForecastListLastEmailMessageTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="email-status-user", password="pass")
        self.org = Organization.objects.create(name="Email Status Org", owner=self.user)
        OrganizationMember.objects.create(
            organization=self.org,
            user=self.user,
            role=OrganizationMember.ROLE_OWNER,
        )
        self.station = Station.objects.create(org=self.org, name="Email Status Station", capacity_mw=1.2)
        self.schedule = ForecastSchedule.objects.create(
            station=self.station,
            enabled=True,
            days=1,
            run_time=timezone.datetime.strptime("06:10", "%H:%M").time(),
            emails="mail@example.com",
            last_email_forecast_date=timezone.datetime(2026, 2, 13).date(),
            last_email_sent_at=timezone.make_aware(timezone.datetime(2026, 2, 12, 6, 11)),
            last_email_status="Прогноз за 13.02.2026 отправлен в 06:11",
        )
        self.client.login(username="email-status-user", password="pass")

    def test_forecast_list_shows_last_email_message(self):
        response = self.client.get(f"/dashboard/station/{self.station.pk}/forecast/list/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Прогноз за 13.02.2026")
        self.assertContains(response, "отправлен в 06:11")


class HistoryDatetimeParsingTests(TestCase):
    def test_prefers_day_first_for_ambiguous_dates(self):
        series = pd.Series(["01/03/2026 10:00:00", "02/03/2026 09:00:00"])

        parsed = _parse_history_datetime(series)

        self.assertEqual(parsed.iloc[0].month, 3)
        self.assertEqual(parsed.iloc[0].day, 1)
        self.assertEqual(parsed.iloc[1].month, 3)
        self.assertEqual(parsed.iloc[1].day, 2)


class StationVisibilityTests(TestCase):
    def setUp(self):
        self.owner = User.objects.create_user(username="org-owner", password="pass")
        self.other = User.objects.create_user(username="other-user", password="pass")
        self.admin = User.objects.create_superuser(username="super", email="super@example.com", password="pass")

        self.owner_org = Organization.objects.create(name="Owner Org", owner=self.owner)
        self.other_org = Organization.objects.create(name="Other Org", owner=self.other)
        self.admin_org = Organization.objects.create(name="Admin Org", owner=self.admin)

        self.owner_station = Station.objects.create(org=self.owner_org, name="Owner Station")
        self.other_station = Station.objects.create(org=self.other_org, name="Other Station")
        self.admin_station = Station.objects.create(org=self.admin_org, name="Admin Station")

    def test_owner_without_membership_sees_own_station(self):
        self.client.login(username="org-owner", password="pass")

        response = self.client.get("/dashboard/")

        self.assertEqual(response.status_code, 200)
        stations = list(response.context["stations"])
        self.assertEqual([st.id for st in stations], [self.owner_station.id])

    def test_superuser_sees_only_owned_or_member_stations(self):
        OrganizationMember.objects.create(
            organization=self.owner_org,
            user=self.admin,
            role=OrganizationMember.ROLE_ADMIN,
        )
        self.client.login(username="super", password="pass")

        response = self.client.get("/dashboard/")

        self.assertEqual(response.status_code, 200)
        station_ids = {st.id for st in response.context["stations"]}
        self.assertEqual(station_ids, {self.admin_station.id, self.owner_station.id})
        self.assertNotIn(self.other_station.id, station_ids)


class StationOrderingTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="order", password="pass")
        self.org = Organization.objects.create(name="Order Org", owner=self.user)
        self.client.login(username="order", password="pass")

        self.station_a = Station.objects.create(org=self.org, name="A")
        self.station_b = Station.objects.create(org=self.org, name="B")
        self.station_c = Station.objects.create(org=self.org, name="C")

    def test_station_list_sorted_by_sort_order(self):
        names = list(
            Station.objects.filter(org=self.org)
            .order_by("sort_order", "id")
            .values_list("name", flat=True)
        )
        self.assertEqual(names, ["A", "B", "C"])

    def test_move_station_down_swaps_with_next_station(self):
        response = self.client.post(f"/dashboard/station/{self.station_a.pk}/move/down/")

        self.assertEqual(response.status_code, 302)
        names = list(
            Station.objects.filter(org=self.org)
            .order_by("sort_order", "id")
            .values_list("name", flat=True)
        )
        self.assertEqual(names, ["B", "A", "C"])

    def test_move_station_up_swaps_with_previous_station(self):
        response = self.client.post(f"/dashboard/station/{self.station_c.pk}/move/up/")

        self.assertEqual(response.status_code, 302)
        names = list(
            Station.objects.filter(org=self.org)
            .order_by("sort_order", "id")
            .values_list("name", flat=True)
        )
        self.assertEqual(names, ["A", "C", "B"])



class StationUploadClearByDateTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="history-clear-user", password="pass")
        self.org = Organization.objects.create(name="History Org", owner=self.user)
        OrganizationMember.objects.create(org=self.org, user=self.user, role=OrganizationMember.ROLE_OWNER)
        self.station = Station.objects.create(name="SES Clear", org=self.org, capacity_mw=1.2)
        self.client.login(username="history-clear-user", password="pass")

    def _record(self, ts: str, power_kw: float) -> SolarRecord:
        aware_ts = timezone.make_aware(datetime.strptime(ts, "%Y-%m-%d %H:%M:%S"), timezone.get_current_timezone())
        return SolarRecord.objects.create(
            station=self.station,
            history_scope=SolarRecord.HISTORY_SCOPE_MAIN,
            timestamp=aware_ts,
            power_kw=power_kw,
        )

    def test_clear_action_deletes_only_records_in_selected_date_range(self):
        keep_before = self._record("2026-03-28 12:00:00", 1.0)
        delete_mid = self._record("2026-03-29 08:00:00", 2.0)
        delete_late = self._record("2026-03-29 18:00:00", 3.0)
        keep_after = self._record("2026-03-30 09:00:00", 4.0)

        response = self.client.post(
            f"/dashboard/station/{self.station.pk}/upload/",
            data={
                "action": "clear",
                "history_scope": "main",
                "from": "2026-03-29",
                "to": "2026-03-29",
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        remaining_ids = list(
            SolarRecord.objects.filter(station=self.station, history_scope=SolarRecord.HISTORY_SCOPE_MAIN)
            .order_by("timestamp")
            .values_list("id", flat=True)
        )
        self.assertEqual(remaining_ids, [keep_before.id, keep_after.id])
        self.assertNotIn(delete_mid.id, remaining_ids)
        self.assertNotIn(delete_late.id, remaining_ids)

    def test_clear_action_without_dates_deletes_full_history_scope(self):
        self._record("2026-03-28 12:00:00", 1.0)
        self._record("2026-03-29 08:00:00", 2.0)

        response = self.client.post(
            f"/dashboard/station/{self.station.pk}/upload/",
            data={"action": "clear", "history_scope": "main"},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(
            SolarRecord.objects.filter(station=self.station, history_scope=SolarRecord.HISTORY_SCOPE_MAIN).exists()
        )


class RunScheduledForecastsCommandTests(TestCase):
    @patch("dashboard.management.commands.run_scheduled_forecasts.importlib.import_module")
    def test_auto_history_safe_helper_returns_zero_on_import_error(self, import_module_mock):
        import_module_mock.side_effect = RuntimeError("boom")

        class DummyStyle:
            @staticmethod
            def WARNING(value):
                return value

        class DummyStdout:
            def __init__(self):
                self.messages = []

            def write(self, message):
                self.messages.append(message)

        stdout = DummyStdout()
        result = _run_auto_history_updates_safe(stdout, DummyStyle)

        self.assertEqual(result, 0)
        self.assertTrue(any("Auto history skipped due to error" in m for m in stdout.messages))

    @patch("dashboard.management.commands.run_scheduled_forecasts.importlib.import_module")
    def test_auto_history_safe_helper_returns_rows_on_success(self, import_module_mock):
        class DummyModule:
            @staticmethod
            def run_auto_history_updates():
                return 7

        import_module_mock.return_value = DummyModule

        class DummyStyle:
            @staticmethod
            def WARNING(value):
                return value

        class DummyStdout:
            def write(self, message):
                return None

        result = _run_auto_history_updates_safe(DummyStdout(), DummyStyle)

        self.assertEqual(result, 7)


class Ses88MwHistoryScriptTests(TestCase):
    def test_build_history_dataframe_parses_kengir_excel_format(self):
        from dashboard.services.history_scripts.ses_8_8mw import build_history_dataframe

        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            xlsx_path = folder / "СЭС Кенгир 10МВт февраль 2026.xlsx"

            raw = pd.DataFrame(
                [
                    ["Время", "Фактическая выработка", "Активная мощность", "Иррадиация", "", "", "Температура воздуха", "Температура PV"],
                    ["08:00", 73, 0.22, 30.7, "", "", -9.1, -9.0],
                    ["08:15", 155, 0.36, 46.9, "", "", -9.1, -8.9],
                    ["scada", 0, 0.0, 0.0, "", "", 0.0, 0.0],
                    ["08:30", 275, 10.5, 65.8, "", "", -9.2, -8.9],
                ]
            )
            with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
                raw.to_excel(writer, sheet_name="26.02.2026", header=False, index=False)

            station = SimpleNamespace(auto_history_folder=str(folder))
            out = build_history_dataframe(station)

            self.assertEqual(len(out), 1)
            self.assertEqual(str(out.iloc[0]["ds"]), "2026-02-26 09:00:00")
            self.assertAlmostEqual(float(out.iloc[0]["power_kw"]), 290.0)
            self.assertGreater(float(out.iloc[0]["irradiation"]), 40.0)


    def test_build_history_dataframe_parses_excel_time_cells(self):
        from dashboard.services.history_scripts.ses_8_8mw import build_history_dataframe

        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            xlsx_path = folder / "СЭС Кенгир 10МВт март 2026.xlsx"

            raw = pd.DataFrame(
                [
                    ["Время", "Фактическая выработка", "Активная мощность", "Иррадиация", "", "", "Температура воздуха", "Температура PV"],
                    [time(8, 0), 73, 0.22, 30.7, "", "", -9.1, -9.0],
                    [time(8, 15), 155, 0.36, 46.9, "", "", -9.1, -8.9],
                ]
            )
            with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
                raw.to_excel(writer, sheet_name="01.03.2026", header=False, index=False)

            station = SimpleNamespace(auto_history_folder=str(folder))
            out = build_history_dataframe(station)

            self.assertEqual(len(out), 1)
            self.assertEqual(str(out.iloc[0]["ds"]), "2026-03-01 09:00:00")
            self.assertAlmostEqual(float(out.iloc[0]["power_kw"]), 290.0)





class SesShieli20MwHistoryScriptTests(TestCase):
    def test_build_history_dataframe_parses_daily_report_and_filters_noise(self):
        from openpyxl import Workbook
        from dashboard.services.history_scripts.ses_shieli_20mw import build_history_dataframe

        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            xlsx_path = folder / "shieli_report.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["time", "x", "power"])
            ws.append(["05.04.2026 00:00", "", 5])
            ws.append(["05.04.2026 01:00", "", 14])
            ws.append(["05.04.2026 02:00", "", 120])
            ws.append(["05.04.2026 03:00", "", 9])
            wb.save(xlsx_path)

            station = SimpleNamespace(auto_history_folder=str(folder), data_shift_hours=0)
            out = build_history_dataframe(station)

        self.assertEqual(len(out), 2)
        self.assertEqual(str(out.iloc[0]["ds"]), "2026-04-05 05:00:00")
        self.assertEqual(str(out.iloc[1]["ds"]), "2026-04-05 06:00:00")
        self.assertAlmostEqual(float(out.iloc[0]["power_kw"]), 14.0)
        self.assertAlmostEqual(float(out.iloc[1]["power_kw"]), 120.0)

    def test_build_history_dataframe_applies_station_shift(self):
        from openpyxl import Workbook
        from dashboard.services.history_scripts.ses_shieli_20mw import build_history_dataframe

        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            xlsx_path = folder / "shieli_shifted.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["time", "x", "power"])
            ws.append(["06.04.2026 08:00", "", 50])
            wb.save(xlsx_path)

            station = SimpleNamespace(auto_history_folder=str(folder), data_shift_hours=-1)
            out = build_history_dataframe(station)

        self.assertEqual(len(out), 1)
        self.assertEqual(str(out.iloc[0]["ds"]), "2026-04-06 11:00:00")
        self.assertAlmostEqual(float(out.iloc[0]["power_kw"]), 50.0)




class SesSheleAliasHistoryScriptTests(TestCase):
    def test_alias_module_exports_same_builder(self):
        from dashboard.services.history_scripts.ses_shele_20mw import build_history_dataframe as alias_builder
        from dashboard.services.history_scripts.ses_shieli_20mw import build_history_dataframe as base_builder

        self.assertIs(alias_builder, base_builder)


class Ses50BalkhashHistoryScriptTests(TestCase):
    def test_build_history_dataframe_uses_hourly_mean_for_power(self):
        from openpyxl import Workbook
        from dashboard.services.history_scripts.ses_50_balkhash import build_history_dataframe

        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            xlsx_path = folder / "balkhash_report.xlsx"

            wb = Workbook()
            ws = wb.active
            ws["A1"] = "Отчет №01 / 09.03.2026 0:00:00"
            ws.append([])
            ws.append([])
            ws.append(["Время", "Мощность актив", "Иррадиация", "Температура воздуха", "", "", "Температура ФЭМ"])
            ws.append(["09.03 - 08:10", 0.2, 100, 5, "", "", 7])
            ws.append(["09.03 - 08:40", 0.4, 120, 6, "", "", 8])
            wb.save(xlsx_path)

            station = SimpleNamespace(auto_history_folder=str(folder))
            out = build_history_dataframe(station)

        self.assertEqual(len(out), 1)
        self.assertEqual(str(out.iloc[0]["ds"]), "2026-03-09 09:00:00")
        self.assertAlmostEqual(float(out.iloc[0]["power_kw"]), 300.0)

    def test_build_history_dataframe_extracts_year_from_sheet_header(self):
        from openpyxl import Workbook
        from dashboard.services.history_scripts.ses_50_balkhash import build_history_dataframe

        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            xlsx_path = folder / "report_without_year_in_name.xlsx"

            wb = Workbook()
            ws = wb.active
            ws["A1"] = "Отчет №55 / 07.03.2026 0:00:00"
            ws.append([])
            ws.append([])
            ws.append(["Время", "Мощность актив", "Иррадиация", "Температура воздуха", "", "", "Температура ФЭМ"])
            ws.append(["07.03 - 23:15", 0.1, 10, -2, "", "", -1])
            wb.save(xlsx_path)

            station = SimpleNamespace(auto_history_folder=str(folder))
            out = build_history_dataframe(station)

        self.assertEqual(len(out), 1)
        self.assertEqual(str(out.iloc[0]["ds"]), "2026-03-08 00:00:00")

    def test_build_history_dataframe_parses_comma_decimal_text_values(self):
        from openpyxl import Workbook
        from dashboard.services.history_scripts.ses_50_balkhash import build_history_dataframe

        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            xlsx_path = folder / "report_text_numbers.xlsx"

            wb = Workbook()
            ws = wb.active
            ws["A1"] = "Отчет №99 / 11.03.2026 0:00:00"
            ws.append([])
            ws.append([])
            ws.append(["Время", "Мощность актив", "Иррадиация", "Температура воздуха", "", "", "Температура ФЭМ"])
            ws.append(["11.03 - 09:00", "32,18", "396,43", "-15,70", "", "", "-3,67"])
            ws.append(["11.03 - 09:15", "35,02", "438,08", "-15,70", "", "", "-1,67"])
            wb.save(xlsx_path)

            station = SimpleNamespace(auto_history_folder=str(folder))
            out = build_history_dataframe(station)

        self.assertEqual(len(out), 1)
        self.assertEqual(str(out.iloc[0]["ds"]), "2026-03-11 10:00:00")
        self.assertAlmostEqual(float(out.iloc[0]["irradiation"]), 417.255, places=3)
        self.assertAlmostEqual(float(out.iloc[0]["power_kw"]), 33600.0, places=2)

class Ses12MwHistoryScriptTests(TestCase):
    @patch("dashboard.services.history_autofill.collect_share_history_dataframe")
    def test_build_history_dataframe_keeps_share_merge_hours_for_d222_and_report_files(self, collect_mock):
        from dashboard.services.history_scripts.ses_1_2mw import build_history_dataframe

        collect_mock.return_value = pd.DataFrame(
            [
                {
                    "ds": pd.Timestamp("2026-03-12 06:00:00"),
                    "irradiation": 123.4,
                    "air_temp": 11.0,
                    "pv_temp": 16.0,
                    "power_kw": 77.7,
                }
            ]
        )

        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            (folder / "D222152_20260312_0000.csv.gz").write_text("dummy")
            (folder / "reportSPP_JezSolar 1.2 MW_12-03-2026_Plant Statistics Report_by Time.xlsx").write_text("dummy")

            station = SimpleNamespace(auto_history_folder=str(folder))
            out = build_history_dataframe(station)

        self.assertEqual(len(out), 1)
        self.assertEqual(str(out.iloc[0]["ds"]), "2026-03-12 06:00:00")
        self.assertAlmostEqual(float(out.iloc[0]["power_kw"]), 77.7)
        collect_mock.assert_called_once()

    def test_build_history_dataframe_parses_standard_csv_columns(self):
        from dashboard.services.history_scripts.ses_1_2mw import build_history_dataframe

        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            pd.DataFrame(
                [
                    {
                        "ds": "2026-02-26 08:10:00",
                        "Irradiation": 30.7,
                        "Air_Temp": -9.1,
                        "PV_Temp": -9.0,
                        "Power_KW": 220.123,
                    },
                    {
                        "ds": "2026-02-26 08:40:00",
                        "Irradiation": 46.9,
                        "Air_Temp": -9.1,
                        "PV_Temp": -8.9,
                        "Power_KW": 70.222,
                    },
                ]
            ).to_csv(folder / "history_1_2.csv", index=False)

            station = SimpleNamespace(auto_history_folder=str(folder))
            out = build_history_dataframe(station)

            self.assertEqual(len(out), 1)
            self.assertEqual(str(out.iloc[0]["ds"]), "2026-02-26 08:00:00")
            self.assertAlmostEqual(float(out.iloc[0]["power_kw"]), 70.22)

    def test_build_history_dataframe_supports_timestamp_alias(self):
        from dashboard.services.history_scripts.ses_1_2mw import build_history_dataframe

        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            pd.DataFrame(
                [
                    {
                        "timestamp": "2026-03-01 10:15:00",
                        "Irradiation": 120,
                        "Air_Temp": 11,
                        "PV_Temp": 16,
                        "Power_KW": 0,
                    },
                    {
                        "timestamp": "2026-03-01 11:05:00",
                        "Irradiation": 220,
                        "Air_Temp": 12,
                        "PV_Temp": 17,
                        "Power_KW": 130,
                    },
                ]
            ).to_csv(folder / "history_1_2.csv", index=False)

            station = SimpleNamespace(auto_history_folder=str(folder))
            out = build_history_dataframe(station)

            self.assertEqual(len(out), 1)
            self.assertEqual(str(out.iloc[0]["ds"]), "2026-03-01 11:00:00")
            self.assertAlmostEqual(float(out.iloc[0]["power_kw"]), 130.0)



class InprocessSchedulerBootstrapTests(TestCase):
    @patch("dashboard.services.inprocess_scheduler.threading.Thread")
    @patch("dashboard.services.inprocess_scheduler._background_scheduler_enabled", return_value=True)
    def test_start_background_scheduler_starts_daemon_thread_once(self, enabled_mock, thread_cls):
        from dashboard.services import inprocess_scheduler

        inprocess_scheduler._BACKGROUND_THREAD = None

        thread_instance = thread_cls.return_value
        thread_instance.is_alive.return_value = True

        inprocess_scheduler.start_background_scheduler()
        inprocess_scheduler.start_background_scheduler()

        thread_cls.assert_called_once()
        thread_instance.start.assert_called_once_with()
        enabled_mock.assert_called()

    @patch("dashboard.services.inprocess_scheduler._background_scheduler_enabled", return_value=False)
    def test_start_background_scheduler_skips_when_disabled(self, _enabled):
        from dashboard.services import inprocess_scheduler

        inprocess_scheduler._BACKGROUND_THREAD = None

        with patch("dashboard.services.inprocess_scheduler.threading.Thread") as thread_cls:
            inprocess_scheduler.start_background_scheduler()

        thread_cls.assert_not_called()

    @patch("dashboard.services.inprocess_scheduler.run_scheduled_forecasts", return_value=2)
    @patch("dashboard.services.inprocess_scheduler.run_auto_history_updates", return_value=5)
    def test_tick_with_lock_runs_both_jobs(self, history_mock, forecast_mock):
        from dashboard.services.inprocess_scheduler import _run_tick_with_file_lock

        ok = _run_tick_with_file_lock()

        self.assertTrue(ok)
        history_mock.assert_called_once_with()
        forecast_mock.assert_called_once_with()


class ForecastReportFilenameTests(TestCase):
    def test_report_filename_includes_station_name(self):
        from dashboard.services.forecast_reports import build_forecast_report

        user = User.objects.create_user(username="report-file", password="pass")
        org = Organization.objects.create(name="Report Org", owner=user)
        station = Station.objects.create(org=org, name="SES 10 MW")

        report = build_forecast_report(
            station=station,
            days=1,
            weather_source="stub",
            recipients=["mail@example.com"],
            forecast_scope="main",
        )

        self.assertIn("forecast_SES_10_MW_", report.file.name)
        self.assertTrue(report.file.name.endswith("_mw.xlsx"))


class StationListScopeTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="scope_user", password="secret123")
        self.org = Organization.objects.create(name="Scope Org", owner=self.user)
        self.solar_station = Station.objects.create(
            org=self.org,
            name="Solar A",
            station_kind=Station.KIND_SOLAR,
            capacity_mw=10.0,
        )
        self.wind_station = Station.objects.create(
            org=self.org,
            name="Wind A",
            station_kind=Station.KIND_WIND,
            capacity_mw=12.0,
        )

    def test_dashboard_station_list_shows_only_solar(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse("dashboard:station-list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Solar A")
        self.assertNotContains(response, "Wind A")

    def test_wind_station_cannot_be_moved_from_dashboard_endpoint(self):
        self.client.force_login(self.user)

        response = self.client.post(reverse("dashboard:station-move", args=[self.wind_station.pk, "up"]))

        self.assertEqual(response.status_code, 404)

class IrradiationHistoryTrainingTests(TestCase):
    def setUp(self):
        self.owner = User.objects.create_user(username="train-irr-owner", password="pass12345")
        self.org = Organization.objects.create(name="Train Irr Org", owner=self.owner)

    def test_training_uses_ghi_when_both_split_columns_exist(self):
        from dashboard.services.train_models import get_history_dataframe

        station = Station.objects.create(org=self.org, name="Tracker", mount_type=Station.MOUNT_SINGLE_AXIS_TRACKER)
        SolarRecord.objects.create(
            station=station,
            timestamp="2026-01-01T10:00:00+05:00",
            power_kw=100.0,
            irradiation=999.0,
            irradiation_ghi=500.0,
            irradiation_poa=700.0,
            air_temp=20.0,
            pv_temp=25.0,
        )

        df = get_history_dataframe(station)
        self.assertEqual(float(df.iloc[0]["Irradiation"]), 500.0)
        self.assertEqual(float(df.iloc[0]["Irradiation_POA"]), 700.0)

    def test_training_does_not_treat_legacy_poa_as_ghi(self):
        from dashboard.services.train_models import get_history_dataframe

        station = Station.objects.create(
            org=self.org,
            name="POA legacy",
            irradiation_type=Station.IRRADIATION_POA,
        )
        SolarRecord.objects.create(
            station=station,
            timestamp="2026-01-01T10:00:00+05:00",
            power_kw=100.0,
            irradiation=700.0,
            air_temp=20.0,
            pv_temp=25.0,
        )

        df = get_history_dataframe(station)
        self.assertTrue(df.empty)
