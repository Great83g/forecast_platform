from __future__ import annotations

import re
from datetime import datetime
from datetime import time as dt_time
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
HEADER_SCAN_ROWS = 20000
MIN_POWER_KW = 0.0001


DATE_TIME_RE = re.compile(r"^(\d{1,2})\.(\d{1,2})(?:\.(\d{2,4}))?\s*-?\s*(\d{1,2}):(\d{1,2})$")
DAY_ONLY_RE = re.compile(r"^(\d{1,2})\.(\d{1,2})(?:\.(\d{2,4}))?$")
TIME_ONLY_RE = re.compile(r"^(\d{1,2}):(\d{1,2})(?::\d{1,2})?$")


REQUIRED_COLUMNS = {
    "time": ["время"],
    "power": ["мощность актив", "мощность"],
    "irradiation": ["иррадиа"],
    "air_temp": ["температура воздуха", "темп воздуха"],
    "pv_temp": ["температура фэм", "температура модуля", "температура панели"],
}


def _empty_df() -> pd.DataFrame:
    return pd.DataFrame(columns=["ds", "irradiation", "air_temp", "pv_temp", "power_kw"])


def _normalize_text(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").replace("\xa0", " ").strip()


def _extract_year_candidates(*values: str) -> list[int]:
    years: list[int] = []
    for value in values:
        if not value:
            continue
        for match in re.findall(r"\b(20\d{2})\b", value):
            years.append(int(match))
    return years


def _guess_year_for_file(file_path: Path) -> int | None:
    candidates = _extract_year_candidates(file_path.name, str(file_path.parent), str(file_path.parent.parent))

    folder_tokens = [file_path.parent.name, file_path.parent.parent.name]
    for token in folder_tokens:
        if token.isdigit() and len(token) == 2:
            candidates.append(2000 + int(token))

    return max(candidates) if candidates else None


def _detect_header_row(ws) -> tuple[int, dict[str, int]]:
    max_scan_rows = min(max(int(getattr(ws, "max_row", 0) or 0), 1), HEADER_SCAN_ROWS)
    for row_num in range(1, max_scan_rows + 1):
        row_vals = [cell.value for cell in ws[row_num]]
        headers = [_normalize_text(x) for x in row_vals]

        if not any(headers):
            continue

        resolved: dict[str, int] = {}
        for idx, col in enumerate(headers):
            low = col.lower()
            for key, candidates in REQUIRED_COLUMNS.items():
                if key in resolved:
                    continue
                if any(c.lower() in low for c in candidates):
                    resolved[key] = idx

        if all(k in resolved for k in REQUIRED_COLUMNS):
            return row_num, resolved

    # Фолбэк для типового формата отчётов СЭС Балхаш:
    # A=Время, B=Мощность актив., C=Иррадиация, D=Температура воздуха, G=Температура ФЭМ.
    # Бывает, что в файле повреждена/смещена шапка, поэтому используем фиксированные индексы.
    # Запускаем парсинг по всему листу (min_row=1), чтобы поймать блоки отчёта,
    # которые начинаются после тысяч служебных строк.
    return 0, {"time": 0, "power": 1, "irradiation": 2, "air_temp": 3, "pv_temp": 6}


def _parse_ds(value, fallback_year: int | None) -> pd.Timestamp | None:
    if isinstance(value, pd.Timestamp):
        ts = value
        if pd.isna(ts):
            return None
        if ts.year >= 2000:
            return ts
        if fallback_year is None:
            return None
        return pd.Timestamp(
            year=fallback_year,
            month=ts.month,
            day=ts.day,
            hour=ts.hour,
            minute=ts.minute,
            second=ts.second,
        )

    if isinstance(value, datetime):
        ts = pd.Timestamp(value)
        if ts.year >= 2000:
            return ts
        if fallback_year is None:
            return None
        return pd.Timestamp(
            year=fallback_year,
            month=ts.month,
            day=ts.day,
            hour=ts.hour,
            minute=ts.minute,
            second=ts.second,
        )

    if isinstance(value, (int, float)):
        ts = pd.to_datetime(value, unit="D", origin="1899-12-30", errors="coerce")
        if pd.notna(ts):
            ts = pd.Timestamp(ts)
            if ts.year >= 2000:
                return ts
            if fallback_year is not None:
                return pd.Timestamp(
                    year=fallback_year,
                    month=ts.month,
                    day=ts.day,
                    hour=ts.hour,
                    minute=ts.minute,
                    second=ts.second,
                )

    text = _normalize_text(value)
    if not text:
        return None

    m = DATE_TIME_RE.match(text)
    if not m:
        return None

    day, month, year_part, hour, minute = m.groups()
    if year_part:
        year = int(year_part)
        if year < 100:
            year += 2000
    else:
        year = fallback_year
    if year is None:
        return None

    try:
        return pd.Timestamp(year=year, month=int(month), day=int(day), hour=int(hour), minute=int(minute))
    except ValueError:
        return None


def _parse_day_marker(value, fallback_year: int | None) -> tuple[int, int, int] | None:
    text = _normalize_text(value)
    if not text:
        return None

    m = DAY_ONLY_RE.match(text)
    if not m:
        return None

    day, month, year_part = m.groups()
    if year_part:
        year = int(year_part)
        if year < 100:
            year += 2000
    else:
        year = fallback_year

    if year is None:
        return None

    try:
        pd.Timestamp(year=year, month=int(month), day=int(day))
    except ValueError:
        return None
    return int(year), int(month), int(day)


def _parse_time_only(value, current_day: tuple[int, int, int] | None) -> pd.Timestamp | None:
    if current_day is None:
        return None

    if isinstance(value, dt_time):
        y, m, d = current_day
        return pd.Timestamp(year=y, month=m, day=d, hour=value.hour, minute=value.minute, second=value.second)

    text = _normalize_text(value)
    if not text:
        return None

    m = TIME_ONLY_RE.match(text)
    if not m:
        return None

    hour, minute = map(int, m.groups())
    y, mon, d = current_day
    try:
        return pd.Timestamp(year=y, month=mon, day=d, hour=hour, minute=minute)
    except ValueError:
        return None


def _process_one_file(file_path: Path, fallback_year: int | None) -> pd.DataFrame:
    wb = load_workbook(file_path, data_only=True)

    rows = []
    for ws in wb.worksheets:
        try:
            header_row, col_idx = _detect_header_row(ws)
        except Exception:
            continue

        current_day = None
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            time_idx = col_idx["time"]
            if time_idx >= len(row):
                continue

            time_raw = row[time_idx]
            day_marker = _parse_day_marker(time_raw, fallback_year)
            if day_marker is not None:
                current_day = day_marker

            ds = _parse_ds(time_raw, fallback_year)
            if ds is None:
                ds = _parse_time_only(time_raw, current_day)
            if ds is None:
                continue

            rows.append(
                {
                    "ds": ds,
                    "power_raw": row[col_idx["power"]] if col_idx["power"] < len(row) else None,
                    "irradiation": row[col_idx["irradiation"]] if col_idx["irradiation"] < len(row) else None,
                    "air_temp": row[col_idx["air_temp"]] if col_idx["air_temp"] < len(row) else None,
                    "pv_temp": row[col_idx["pv_temp"]] if col_idx["pv_temp"] < len(row) else None,
                }
            )

    if not rows:
        return _empty_df()

    df = pd.DataFrame(rows)
    for col in ["power_raw", "irradiation", "air_temp", "pv_temp"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    df["power_raw"] = df["power_raw"].clip(lower=0)

    hourly = (
        df.set_index("ds")
        .resample("h")
        .agg({
            "irradiation": "mean",
            "air_temp": "mean",
            "pv_temp": "mean",
            "power_raw": "sum",
        })
        .reset_index()
    )

    hourly["power_kw"] = hourly["power_raw"] * 1000.0
    hourly = hourly.drop(columns=["power_raw"])

    hourly = hourly[
        (hourly["power_kw"].fillna(0) > MIN_POWER_KW)
        | (hourly["irradiation"].fillna(0) > 0)
    ].copy()

    for col in ["irradiation", "air_temp", "pv_temp"]:
        hourly[col] = hourly[col].round(3)
    hourly["power_kw"] = hourly["power_kw"].round(3)

    return hourly[["ds", "irradiation", "air_temp", "pv_temp", "power_kw"]]


def _collect_excel_files(folder: Path) -> list[Path]:
    files = [
        p
        for p in folder.rglob("*")
        if p.is_file() and p.suffix.lower() in EXCEL_EXTENSIONS and not p.name.startswith("~$")
    ]
    return sorted(files)


def build_history_dataframe(station) -> pd.DataFrame:
    folder = Path(getattr(station, "auto_history_folder", "") or "")
    if not folder.exists():
        return _empty_df()

    files = _collect_excel_files(folder)
    if not files:
        return _empty_df()

    parts: list[pd.DataFrame] = []
    for file_path in files:
        fallback_year = _guess_year_for_file(file_path)
        try:
            part = _process_one_file(file_path, fallback_year)
            if not part.empty:
                parts.append(part)
        except Exception:
            continue

    if not parts:
        return _empty_df()

    out = pd.concat(parts, ignore_index=True).sort_values("ds").reset_index(drop=True)
    out = out.drop_duplicates(subset=["ds"], keep="last")
    return out[["ds", "irradiation", "air_temp", "pv_temp", "power_kw"]].reset_index(drop=True)
