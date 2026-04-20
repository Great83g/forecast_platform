from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
import pandas as pd

BASE_TIME_SHIFT_HOURS = 4
ABS_MIN_POWER_KW = 10.0
DAY_PEAK_PERCENT = 0.01



def _empty_df() -> pd.DataFrame:
    return pd.DataFrame(columns=["ds", "irradiation", "air_temp", "pv_temp", "power_kw"])



def _norm(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").replace("\xa0", " ").strip()



def _to_float(value) -> float:
    try:
        return float(str(value).replace(" ", "").replace(",", "."))
    except Exception:
        return 0.0



def _parse_dt(value):
    try:
        return pd.to_datetime(value, dayfirst=True)
    except Exception:
        return None



def _station_shift_hours(station) -> int:
    raw = getattr(station, "data_shift_hours", 0)
    try:
        return int(raw)
    except Exception:
        return 0



def _process_file(file_path: Path, shift_hours: int) -> pd.DataFrame:
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    rows: list[dict] = []
    for r in range(1, ws.max_row + 1):
        token = _norm(ws.cell(r, 1).value)
        if "." not in token or ":" not in token:
            continue

        dt = _parse_dt(token)
        if dt is None:
            continue

        power = _to_float(ws.cell(r, 3).value)
        dt = dt + pd.Timedelta(hours=shift_hours)
        rows.append(
            {
                "ds": dt,
                "irradiation": pd.NA,
                "air_temp": pd.NA,
                "pv_temp": pd.NA,
                "power_kw": power,
            }
        )

    if not rows:
        return _empty_df()

    return pd.DataFrame(rows)



def _clean_by_day(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return _empty_df()

    work = df.copy()
    work["_date"] = work["ds"].dt.date
    clean_days: list[pd.DataFrame] = []

    for _, group in work.groupby("_date"):
        g = group.sort_values("ds").copy()
        g_nonzero = g[g["power_kw"] > 0].copy()
        if g_nonzero.empty:
            continue

        day_max = g_nonzero["power_kw"].max()
        if day_max <= 0:
            continue

        noise_threshold = max(ABS_MIN_POWER_KW, day_max * DAY_PEAK_PERCENT)
        g_signal = g_nonzero[g_nonzero["power_kw"] >= noise_threshold].copy()
        if g_signal.empty:
            continue

        start_dt = g_signal["ds"].min()
        end_dt = g_signal["ds"].max()

        g_daylight = g[(g["ds"] >= start_dt) & (g["ds"] <= end_dt)].copy()
        g_daylight = g_daylight[g_daylight["power_kw"] >= ABS_MIN_POWER_KW].copy()
        if not g_daylight.empty:
            clean_days.append(g_daylight)

    if not clean_days:
        return _empty_df()

    out = pd.concat(clean_days, ignore_index=True)
    out = out.sort_values("ds").drop_duplicates(subset=["ds"], keep="last").reset_index(drop=True)
    out["power_kw"] = pd.to_numeric(out["power_kw"], errors="coerce").round(2)

    return out[["ds", "irradiation", "air_temp", "pv_temp", "power_kw"]]



def build_history_dataframe(station) -> pd.DataFrame:
    folder = Path(getattr(station, "auto_history_folder", "") or "")
    if not folder.exists():
        return _empty_df()

    shift_hours = BASE_TIME_SHIFT_HOURS + _station_shift_hours(station)

    files = sorted(p for p in folder.rglob("*") if p.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"} and not p.name.startswith("~$"))
    if not files:
        return _empty_df()

    parts: list[pd.DataFrame] = []
    for file_path in files:
        try:
            parsed = _process_file(file_path, shift_hours)
        except Exception:
            continue
        if not parsed.empty:
            parts.append(parsed)

    if not parts:
        return _empty_df()

    raw = pd.concat(parts, ignore_index=True)
    raw = raw.sort_values("ds").drop_duplicates(subset=["ds"], keep="last").reset_index(drop=True)
    return _clean_by_day(raw)
